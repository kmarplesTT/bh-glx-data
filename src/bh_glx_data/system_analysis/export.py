"""Export module for system analysis.

This module provides Excel export functionality for query results with
formatting and heatmap visualization.
"""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Union

from bh_glx_data.core.exceptions import ExcelGenerationError
from bh_glx_data.system_analysis.database import DatabaseManager, DatabaseStats
from bh_glx_data.system_analysis.excel_formatters import (
    create_or_open_workbook,
    generate_unique_worksheet_name,
    write_ber_plot_to_worksheet,
    write_heatmap_to_worksheet,
    write_histogram_to_worksheet,
    write_table_to_worksheet,
)
from bh_glx_data.system_analysis.query_engine import (
    AggregatedHostStats,
    BERHistogram,
    BERPlot,
    BERStatistics,
    CustomThresholdCounts,
    ThresholdExceededCounts,
    TrainingFailureCounts,
)
from bh_glx_data.system_analysis.visualization import (
    BER_COLOR_SCHEMES,
    ColorScheme,
    COUNT_COLOR_SCHEMES,
)

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportResult:
    """Result of an Excel export operation.

    Attributes:
        output_path: Path to output file
        worksheet_name: Name of worksheet created
        rows_written: Number of data rows written
        file_existed: Whether file existed before export
    """

    output_path: Path
    worksheet_name: str
    rows_written: int
    file_existed: bool


def sanitize_worksheet_name(name: str) -> str:
    r"""Sanitize worksheet name by replacing invalid Excel characters.

    Excel doesn't allow these characters in worksheet names: : \ / ? * [ ]

    Args:
        name: Original worksheet name

    Returns:
        Sanitized worksheet name
    """
    # Replace invalid characters with underscore
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, "_")
    return sanitized


class ExcelExporter:
    """Export query results to Excel files with formatting.

    This class handles exporting query results (stats, counts, histograms)
    to Excel with tables and heatmaps, appending to existing files or
    creating new ones.

    Attributes:
        db: DatabaseManager instance
    """

    def __init__(self, db_manager: DatabaseManager):
        """Initialize Excel exporter.

        Args:
            db_manager: DatabaseManager instance
        """
        self.db = db_manager

    def export_ber_statistics(
        self,
        stats: BERStatistics,
        output_path: Path,
        lane_spec: str,
        format: str = "table",
        color_scheme: Optional[ColorScheme] = None,
        statistic: str = "max",
    ) -> ExcelExportResult:
        """Export BER statistics to Excel (table or heatmap format).

        Args:
            stats: BER statistics result from query
            output_path: Path to Excel file
            lane_spec: Lane specification string for worksheet name
            format: "table" or "heatmap"
            color_scheme: Optional color scheme for heatmap
            statistic: Statistic to display (avg, min, max, high_ber) - for heatmap

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Generate worksheet name: "Stats - <lane_spec>"
            base_name = f"Stats - {lane_spec}"
            base_name = sanitize_worksheet_name(base_name)
            ws_name = generate_unique_worksheet_name(wb, base_name)
            ws = wb.create_sheet(ws_name)

            # Prepare metadata
            metadata = {
                "Total Samples": stats.num_tests,
                "Unique Systems": stats.num_systems,
                "Train Speeds": list(stats.train_speeds) if stats.train_speeds else "All",
            }

            if format == "table":
                # Prepare table data
                data = self._prepare_ber_statistics_table_data(stats)
                headers = ["bus_id", "eth_id", "lane", "Min BER", "Avg BER", "Max BER", "High BER Count", "Samples"]

                # Define number formats for BER columns (scientific notation with 2 decimal places)
                column_formats = {
                    "Min BER": "0.00E+00",
                    "Avg BER": "0.00E+00",
                    "Max BER": "0.00E+00",
                }

                write_table_to_worksheet(
                    ws,
                    data,
                    headers,
                    title=f"BER Statistics - {lane_spec}",
                    metadata=metadata,
                    column_formats=column_formats,
                )
                rows_written = len(data["bus_id"])

            else:  # heatmap
                # For heatmap, use the specified statistic
                lane_data = {}
                for lane_id, lane_stat in stats.lane_stats.items():
                    if statistic == "avg":
                        lane_data[lane_id] = lane_stat.avg_ber
                    elif statistic == "min":
                        lane_data[lane_id] = lane_stat.min_ber
                    elif statistic == "high_ber":
                        lane_data[lane_id] = lane_stat.high_ber_count
                    else:  # max (default)
                        lane_data[lane_id] = lane_stat.max_ber

                # Add statistic name to metadata for display under heading
                metadata_with_stat = {**metadata, "Statistic": statistic.upper()}

                write_heatmap_to_worksheet(
                    ws,
                    lane_data,
                    color_scheme or BER_COLOR_SCHEMES["default"],  # Use default if none provided
                    is_ber_metric=(statistic != "high_ber"),  # high_ber is count metric
                    title=f"BER Statistics Heatmap - {lane_spec}",
                    metadata=metadata_with_stat,
                )
                rows_written = len(lane_data)

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export BER statistics: {e}", output_path=str(output_path)
            ) from e

    def export_count_data(
        self,
        counts: Union[ThresholdExceededCounts, CustomThresholdCounts, TrainingFailureCounts],
        output_path: Path,
        lane_spec: str,
        format: str = "table",
        color_scheme: Optional[ColorScheme] = None,
    ) -> ExcelExportResult:
        """Export count data to Excel (table or heatmap format).

        Args:
            counts: Count data result from query
            output_path: Path to Excel file
            lane_spec: Lane specification string for worksheet name
            format: "table" or "heatmap"
            color_scheme: Optional color scheme for heatmap

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Determine command type for worksheet name
            if isinstance(counts, ThresholdExceededCounts):
                cmd_type = "Threshold"
            elif isinstance(counts, CustomThresholdCounts):
                cmd_type = "Custom"
            elif isinstance(counts, TrainingFailureCounts):
                cmd_type = "Training"
            else:
                cmd_type = "Counts"

            # Generate worksheet name: "<command> - <format>"
            base_name = f"{cmd_type} - {format}"
            ws_name = generate_unique_worksheet_name(wb, base_name)
            ws = wb.create_sheet(ws_name)

            # Prepare metadata
            metadata = {
                "Total Count": counts.num_tests,
                "Unique Systems": counts.num_systems,
                "Train Speeds": list(counts.train_speeds) if counts.train_speeds else "All",
            }

            # Add threshold to metadata for custom command (will appear under heading)
            if isinstance(counts, CustomThresholdCounts):
                metadata["BER Threshold"] = counts.threshold  # Store as number for proper formatting

            if format == "table":
                # Prepare table data with separate columns for bus_id, eth_id, lane
                data = self._prepare_count_table_data(counts)
                headers = ["bus_id", "eth_id", "lane", "Count"]
                write_table_to_worksheet(
                    ws, data, headers, title=f"{cmd_type} Counts - {lane_spec}", metadata=metadata
                )
                rows_written = len(data["bus_id"])

            else:  # heatmap
                write_heatmap_to_worksheet(
                    ws,
                    counts.lane_counts,
                    color_scheme or COUNT_COLOR_SCHEMES["default"],
                    is_ber_metric=False,
                    title=f"{cmd_type} Counts Heatmap - {lane_spec}",
                    metadata=metadata,
                )
                rows_written = len(counts.lane_counts)

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export count data: {e}", output_path=str(output_path)
            ) from e

    def export_histogram(
        self,
        histogram: Union[BERHistogram, List[BERHistogram]],
        output_path: Path,
        lane_spec: str,
    ) -> ExcelExportResult:
        """Export histogram(s) to Excel with column chart.

        Args:
            histogram: BERHistogram or list of histograms
            output_path: Path to Excel file
            lane_spec: Lane specification string for worksheet name

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Generate worksheet name: "Hist - <lane_spec>"
            base_name = f"Hist - {lane_spec}"
            base_name = sanitize_worksheet_name(base_name)
            ws_name = generate_unique_worksheet_name(wb, base_name)
            ws = wb.create_sheet(ws_name)

            # Normalize to list
            histograms = [histogram] if isinstance(histogram, BERHistogram) else histogram

            # Prepare metadata
            total_samples = sum(sum(count for _, count in h.bins) for h in histograms)
            unique_systems = len(
                set(h.lane_id.split("/")[0] for h in histograms if "/" in h.lane_id)
            )

            metadata = {
                "Total Samples": total_samples,
                "Unique Systems": unique_systems if unique_systems > 0 else "N/A",
                "Lanes Analyzed": len(histograms),
            }

            # Check if train_speeds available (from first histogram)
            if histograms and hasattr(histograms[0], "train_speeds"):
                metadata["Train Speeds"] = list(histograms[0].train_speeds)

            # Write histogram
            write_histogram_to_worksheet(
                ws, histograms, title=f"BER Distribution - {lane_spec}", metadata=metadata
            )

            rows_written = sum(len(h.bins) for h in histograms)

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export histogram: {e}", output_path=str(output_path)
            ) from e

    def export_advanced_stats(
        self,
        stats: Union[AggregatedHostStats, List[AggregatedHostStats]],
        output_path: Path,
        lane_spec: str,
        color_scheme: Optional[ColorScheme] = None,
    ) -> ExcelExportResult:
        """Export advanced statistics (two tables per lane).

        Args:
            stats: AggregatedHostStats or list of stats
            output_path: Path to Excel file
            lane_spec: Lane specification string for worksheet name
            color_scheme: Optional color scheme for BER value cells

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Generate worksheet name: "Adv Stats - <lane_spec>"
            base_name = f"Adv Stats - {lane_spec}"
            base_name = sanitize_worksheet_name(base_name)
            ws_name = generate_unique_worksheet_name(wb, base_name)
            ws = wb.create_sheet(ws_name)

            # Normalize to list
            stats_list = [stats] if isinstance(stats, AggregatedHostStats) else stats

            # Use provided color scheme or default BER color scheme
            effective_color_scheme = color_scheme or BER_COLOR_SCHEMES["default"]

            row = 1
            rows_written = 0

            from openpyxl.styles import Color, Font, PatternFill

            from bh_glx_data.system_analysis.excel_formatters import (
                apply_cell_background_color,
                get_excel_color_for_value,
            )

            for agg_stats in stats_list:
                # Title for this lane
                ws.cell(row, 1, f"Advanced Statistics - {agg_stats.lane_id}")
                ws.cell(row, 1).font = Font(bold=True, size=14)
                row += 2

                # Table 1: Per-Host Statistics

                ws.cell(row, 1, "Per-Host Statistics")
                ws.cell(row, 1).font = Font(bold=True, size=12)
                row += 1

                # Headers
                headers = ["Host", "Min BER", "Avg BER", "Max BER", "Samples"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )
                row += 1

                # Data rows
                for host_stat in sorted(agg_stats.host_stats, key=lambda h: h.host):
                    ws.cell(row, 1, host_stat.host)

                    # Min BER
                    min_cell = ws.cell(row, 2)
                    if host_stat.min_ber is not None:
                        min_cell.value = host_stat.min_ber
                        min_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(host_stat.min_ber, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(min_cell, color)
                    else:
                        min_cell.value = "-"

                    # Avg BER
                    avg_cell = ws.cell(row, 3)
                    if host_stat.avg_ber is not None:
                        avg_cell.value = host_stat.avg_ber
                        avg_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(host_stat.avg_ber, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(avg_cell, color)
                    else:
                        avg_cell.value = "-"

                    # Max BER
                    max_cell = ws.cell(row, 4)
                    if host_stat.max_ber is not None:
                        max_cell.value = host_stat.max_ber
                        max_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(host_stat.max_ber, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(max_cell, color)
                    else:
                        max_cell.value = "-"

                    ws.cell(row, 5, host_stat.sample_count)
                    row += 1
                    rows_written += 1

                row += 2  # Spacing

                # Table 2: Statistics of Host Statistics
                ws.cell(row, 1, "Statistics of Host Statistics")
                ws.cell(row, 1).font = Font(bold=True, size=12)
                row += 1

                # Headers
                headers = ["Metric", "Minimum", "Average", "Maximum"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )
                row += 1

                # Data rows
                stats_data = [
                    ("MIN", agg_stats.min_of_mins, agg_stats.avg_of_mins, agg_stats.max_of_mins),
                    ("AVG", agg_stats.min_of_avgs, agg_stats.avg_of_avgs, agg_stats.max_of_avgs),
                    ("MAX", agg_stats.min_of_maxs, agg_stats.avg_of_maxs, agg_stats.max_of_maxs),
                ]
                for metric, min_val, avg_val, max_val in stats_data:
                    ws.cell(row, 1, metric)

                    # Min value
                    min_cell = ws.cell(row, 2)
                    if min_val is not None:
                        min_cell.value = min_val
                        min_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(min_val, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(min_cell, color)
                    else:
                        min_cell.value = "-"

                    # Avg value
                    avg_cell = ws.cell(row, 3)
                    if avg_val is not None:
                        avg_cell.value = avg_val
                        avg_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(avg_val, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(avg_cell, color)
                    else:
                        avg_cell.value = "-"

                    # Max value
                    max_cell = ws.cell(row, 4)
                    if max_val is not None:
                        max_cell.value = max_val
                        max_cell.number_format = "0.00E+00"
                        # Apply color scheme (background only - use black text)
                        color = get_excel_color_for_value(max_val, effective_color_scheme, is_ber_metric=True)
                        apply_cell_background_color(max_cell, color)
                    else:
                        max_cell.value = "-"

                    row += 1
                    rows_written += 1

                row += 3  # Spacing between lanes

            # Add color legend after all lane tables
            row += 1
            ws.cell(row, 1, "Color Legend:")
            ws.cell(row, 1).font = Font(bold=True)
            row += 1

            # Generate legend from color scheme thresholds
            sorted_thresholds = sorted(effective_color_scheme.thresholds)
            for i, (threshold, color) in enumerate(sorted_thresholds):
                threshold_text = f"<= {threshold:.2e}"
                ws.cell(row, 1, threshold_text)

                from bh_glx_data.system_analysis.excel_formatters import (
                    apply_cell_background_color,
                    map_terminal_color_to_excel,
                )

                color_hex = map_terminal_color_to_excel(color)
                apply_cell_background_color(ws.cell(row, 2, ""), color_hex)
                row += 1

            # Add default color for values exceeding all thresholds
            last_threshold_color = sorted_thresholds[-1][1]
            if effective_color_scheme.default_color != last_threshold_color:
                threshold_text = f"> {sorted_thresholds[-1][0]:.2e}"
                ws.cell(row, 1, threshold_text)

                color_hex = map_terminal_color_to_excel(effective_color_scheme.default_color)
                apply_cell_background_color(ws.cell(row, 2, ""), color_hex)
                row += 1

            # Add metadata summary at bottom
            from bh_glx_data.system_analysis.excel_formatters import write_metadata_section

            row += 2
            total_samples = sum(hs.sample_count for s in stats_list for hs in s.host_stats)
            unique_systems = len(set(hs.host for s in stats_list for hs in s.host_stats))

            metadata = {
                "Total Samples": total_samples,
                "Unique Systems": unique_systems,
            }
            write_metadata_section(ws, row, metadata)

            # Adjust column widths
            ws.column_dimensions["A"].width = 20
            for col in ["B", "C", "D"]:
                ws.column_dimensions[col].width = 15
            ws.column_dimensions["E"].width = 10

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export advanced stats: {e}", output_path=str(output_path)
            ) from e

    def export_database_info(
        self,
        info: DatabaseStats,
        output_path: Path,
    ) -> ExcelExportResult:
        """Export database info to Excel.

        Args:
            info: DatabaseStats with database metadata
            output_path: Path to Excel file

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Generate worksheet name
            ws_name = generate_unique_worksheet_name(wb, "Database Info")
            ws = wb.create_sheet(ws_name)

            # Prepare data - store numbers as numbers, not strings
            # Use three columns for better formatting
            data = {
                "Property": [
                    "Database Path",
                    "Total Samples",
                    "Total Tests",
                    "Unique Systems",
                    "Train Speeds",
                    "Date Range",
                    "",  # Blank row
                    "Status Breakdown:",
                ],
                "Value": [
                    str(self.db.db_path),
                    info.total_samples,  # Store as number
                    info.total_tests,  # Store as number
                    info.unique_hosts,  # Store as number
                    ", ".join(str(s) for s in info.unique_speeds),  # Keep as string (comma-separated)
                    f"{info.date_range[0]} to {info.date_range[1]}" if info.date_range else "N/A",
                    "",
                    "",
                ],
                "Percentage": [
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ],
            }

            # Add status breakdown with counts as numbers and percentages
            for status, count in info.status_breakdown.items():
                percentage = (count / info.total_samples * 100) if info.total_samples > 0 else 0
                data["Property"].append(f"  {status}")
                data["Value"].append(count)  # Store count as number
                data["Percentage"].append(percentage / 100)  # Store as decimal for percentage formatting

            # Add blank row and ingestion count
            data["Property"].extend(["", "Total Ingestions"])
            data["Value"].extend(["", info.total_ingestions])  # Store as number
            data["Percentage"].extend(["", ""])

            headers = ["Property", "Value", "Percentage"]
            column_formats = {
                "Value": "#,##0",  # Thousands separator for numbers
                "Percentage": "0.0%",  # Percentage format
            }
            write_table_to_worksheet(
                ws, data, headers, title="Database Information", column_formats=column_formats
            )

            rows_written = len(data["Property"])

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export database info: {e}", output_path=str(output_path)
            ) from e

    def export_ber_plot(
        self,
        plot: Union[BERPlot, List[BERPlot]],
        output_path: Path,
        lane_spec: str,
    ) -> ExcelExportResult:
        """Export BER plot(s) to Excel with line chart.

        Args:
            plot: BERPlot or list of plots
            output_path: Path to Excel file
            lane_spec: Lane specification string for worksheet name

        Returns:
            ExcelExportResult with export details

        Raises:
            ExcelGenerationError: If export fails
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create or open workbook
            wb, file_existed = create_or_open_workbook(output_path)

            # Generate worksheet name: "Plot - <lane_spec>"
            base_name = f"Plot - {lane_spec}"
            base_name = sanitize_worksheet_name(base_name)
            ws_name = generate_unique_worksheet_name(wb, base_name)
            ws = wb.create_sheet(ws_name)

            # Normalize to list
            plots = [plot] if isinstance(plot, BERPlot) else plot

            # Prepare metadata
            total_points = sum(len(p.data_points) for p in plots)
            unique_systems = len(set(p.lane_id.split("/")[0] for p in plots if "/" in p.lane_id))

            metadata = {
                "Total Data Points": total_points,
                "Unique Systems": unique_systems if unique_systems > 0 else "N/A",
                "Lanes Plotted": len(plots),
            }

            # Check if train_speeds available (from first plot)
            if plots and hasattr(plots[0], "train_speeds"):
                metadata["Train Speeds"] = list(plots[0].train_speeds)

            # Write plot
            write_ber_plot_to_worksheet(
                ws, plots, title=f"BER Plot - {lane_spec}", metadata=metadata
            )

            rows_written = sum(len(p.data_points) for p in plots)

            # Save workbook
            wb.save(output_path)

            return ExcelExportResult(
                output_path=output_path,
                worksheet_name=ws_name,
                rows_written=rows_written,
                file_existed=file_existed,
            )

        except Exception as e:
            raise ExcelGenerationError(
                f"Failed to export BER plot: {e}", output_path=str(output_path)
            ) from e

    def _prepare_ber_statistics_table_data(self, stats: BERStatistics) -> Dict[str, List]:
        """Prepare BER statistics data for table format.

        Args:
            stats: BER statistics

        Returns:
            Dictionary mapping column names to data lists
        """
        data = {
            "bus_id": [],
            "eth_id": [],
            "lane": [],
            "Min BER": [],
            "Avg BER": [],
            "Max BER": [],
            "High BER Count": [],
            "Samples": [],
        }

        for lane_id in sorted(stats.lane_stats.keys()):
            lane_stat = stats.lane_stats[lane_id]

            # Parse lane_id: "hostname/bus_id/eth_id/lane_N" or "bus_id/eth_id/lane_N"
            parts = lane_id.split("/")
            if len(parts) == 4:
                # Format: hostname/bus_id/eth_id/lane_N
                bus_id = parts[1]
                eth_id = parts[2]
                lane_str = parts[3]
            elif len(parts) == 3:
                # Format: bus_id/eth_id/lane_N
                bus_id = parts[0]
                eth_id = parts[1]
                lane_str = parts[2]
            else:
                # Unexpected format, use full lane_id
                bus_id = lane_id
                eth_id = "-"
                lane_str = "-"

            # Extract lane number from "lane_N" or "laneN" format as integer
            if lane_str.startswith("lane_"):
                lane_num = int(lane_str.split("_")[1])
            elif lane_str.startswith("lane"):
                # Handle "laneN" format (no underscore)
                lane_num = int(lane_str.replace("lane", ""))
            else:
                lane_num = int(lane_str) if lane_str.isdigit() else lane_str

            data["bus_id"].append(bus_id)
            data["eth_id"].append(eth_id)
            data["lane"].append(lane_num)

            # Store numeric values (or "-" for None) - formatting will be applied by write_table_to_worksheet
            data["Min BER"].append(lane_stat.min_ber if lane_stat.min_ber is not None else "-")
            data["Avg BER"].append(lane_stat.avg_ber if lane_stat.avg_ber is not None else "-")
            data["Max BER"].append(lane_stat.max_ber if lane_stat.max_ber is not None else "-")
            data["High BER Count"].append(lane_stat.high_ber_count)
            data["Samples"].append(lane_stat.sample_count)

        return data

    def _prepare_count_table_data(
        self, counts: Union[ThresholdExceededCounts, CustomThresholdCounts, TrainingFailureCounts]
    ) -> Dict[str, List]:
        """Prepare count data for table format with separate columns.

        Args:
            counts: Count data

        Returns:
            Dictionary mapping column names to data lists
        """
        data = {
            "bus_id": [],
            "eth_id": [],
            "lane": [],
            "Count": [],
        }

        for lane_id in sorted(counts.lane_counts.keys()):
            count = counts.lane_counts[lane_id]

            # Parse lane_id: "hostname/bus_id/eth_id/lane_N" or "bus_id/eth_id/lane_N"
            parts = lane_id.split("/")
            if len(parts) == 4:
                # Format: hostname/bus_id/eth_id/lane_N
                bus_id = parts[1]
                eth_id = parts[2]
                lane_str = parts[3]
            elif len(parts) == 3:
                # Format: bus_id/eth_id/lane_N
                bus_id = parts[0]
                eth_id = parts[1]
                lane_str = parts[2]
            else:
                # Unexpected format, use full lane_id
                bus_id = lane_id
                eth_id = "-"
                lane_str = "-"

            # Extract lane number from "lane_N" or "laneN" format as integer
            if lane_str.startswith("lane_"):
                lane_num = int(lane_str.split("_")[1])
            elif lane_str.startswith("lane"):
                # Handle "laneN" format (no underscore)
                lane_num = int(lane_str.replace("lane", ""))
            else:
                lane_num = int(lane_str) if lane_str.isdigit() else lane_str

            data["bus_id"].append(bus_id)
            data["eth_id"].append(eth_id)
            data["lane"].append(lane_num)
            data["Count"].append(count)

        return data
