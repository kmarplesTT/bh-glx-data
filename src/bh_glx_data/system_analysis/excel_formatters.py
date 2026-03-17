"""Excel formatting utilities for system analysis.

This module provides utilities for creating and formatting Excel worksheets
with colors, tables, heatmaps, and charts.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bh_glx_data.system_analysis.query_engine import BERHistogram, BERPlot
from bh_glx_data.system_analysis.visualization import ColorScheme

logger = logging.getLogger(__name__)

# Terminal color to Excel hex color mapping (ARGB format with FF alpha prefix)
# Using lighter colors (matching histogram gradient) for better readability with black text
TERMINAL_TO_EXCEL_COLOR_MAP = {
    "color(28)": "FF4CAF50",  # GREEN -> medium green (better with black text)
    "color(106)": "FF9DC34D",  # YELLOW_GREEN -> yellow-green
    "color(184)": "FFFFD700",  # YELLOW -> gold
    "color(172)": "FFFFA500",  # ORANGE -> orange
    "color(124)": "FFDC143C",  # RED -> crimson red
    # Default colors
    "green": "FF4CAF50",
    "yellow": "FFFFD700",
    "orange": "FFFFA500",
    "red": "FFDC143C",
}


def create_or_open_workbook(path: Path) -> Tuple[Workbook, bool]:
    """Create new or open existing Excel workbook.

    Args:
        path: Path to Excel file

    Returns:
        Tuple of (workbook, file_existed)
    """
    if path.exists():
        wb = load_workbook(path)
        return wb, True
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        return wb, False


def generate_unique_worksheet_name(wb: Workbook, base_name: str) -> str:
    """Generate unique worksheet name avoiding collisions.

    Base name should include lane_spec for descriptive names:
    - "Stats - all"
    - "Stats - 01:00.0/ETH07"
    - "Histogram - 01:00.0/*"

    If collision: "Stats - all", "Stats - all (2)", "Stats - all (3)"

    Args:
        wb: Workbook to check for existing names
        base_name: Base name for worksheet

    Returns:
        Unique worksheet name
    """
    # Excel worksheet names must be <= 31 characters
    max_length = 31

    # Truncate base name if needed (leave room for " (N)")
    if len(base_name) > max_length:
        base_name = base_name[: max_length - 5] + "..."

    # Check if base name is available
    if base_name not in wb.sheetnames:
        return base_name

    # Try numbered versions
    counter = 2
    while True:
        candidate = f"{base_name} ({counter})"
        if len(candidate) > max_length:
            # Truncate base name further to fit counter
            truncated_base = base_name[: max_length - len(f" ({counter})") - 3] + "..."
            candidate = f"{truncated_base} ({counter})"

        if candidate not in wb.sheetnames:
            return candidate

        counter += 1
        if counter > 100:  # Safety limit
            raise ValueError(f"Could not generate unique worksheet name for: {base_name}")


def apply_cell_background_color(cell, color: str) -> None:
    """Apply background color to cell (not font color).

    Args:
        cell: openpyxl Cell object
        color: Hex color string (e.g., "006400" or "FF006400") or terminal color name
    """
    # Map terminal colors to Excel hex colors
    if color.startswith("color(") or color in TERMINAL_TO_EXCEL_COLOR_MAP:
        hex_color = map_terminal_color_to_excel(color)
    else:
        hex_color = color

    # Remove '#' prefix if present
    if hex_color.startswith("#"):
        hex_color = hex_color[1:]

    # Ensure we have 8-character ARGB format (add FF alpha prefix if needed)
    if len(hex_color) == 6:
        hex_color = "FF" + hex_color

    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def map_terminal_color_to_excel(terminal_color: str) -> str:
    """Map Rich terminal color codes to Excel hex colors.

    Args:
        terminal_color: Terminal color string (e.g., "color(28)", "green")

    Returns:
        Excel hex color with alpha channel (e.g., "FF006400")
    """
    return TERMINAL_TO_EXCEL_COLOR_MAP.get(terminal_color, "FFFFFFFF")  # Default to white


# Removed should_use_white_font function - all heatmaps now use black text


def get_excel_color_for_value(
    value: Union[float, int, None], color_scheme: ColorScheme, is_ber_metric: bool
) -> str:
    """Get Excel hex color for value based on color scheme.

    Thresholds define upper bounds (exclusive). For example:
    thresholds=[(1e-12, "green"), (1e-7, "yellow"), (1e-6, "bright_yellow")]
    means:
    - value <= 1e-12: green
    - 1e-12 < value <= 1e-7: yellow
    - 1e-7 < value <= 1e-6: bright_yellow
    - value > 1e-6: default_color

    Args:
        value: Value to get color for (or None)
        color_scheme: Color scheme with thresholds
        is_ber_metric: True if BER metric, False if count metric (unused, kept for compatibility)

    Returns:
        Excel hex color with alpha channel (e.g., "FF006400")
    """
    if value is None:
        return "FFFFFFFF"  # White for None/null values

    # Sort thresholds in ascending order
    sorted_thresholds = sorted(color_scheme.thresholds)

    # Find appropriate color based on thresholds
    for threshold, terminal_color in sorted_thresholds:
        if value <= threshold:
            return map_terminal_color_to_excel(terminal_color)

    # Value exceeds all thresholds - use default color
    return map_terminal_color_to_excel(color_scheme.default_color)


def write_metadata_section(ws: Worksheet, start_row: int, metadata: Dict[str, Any]) -> int:
    """Write metadata summary section to worksheet.

    Args:
        ws: Worksheet to write to
        start_row: Starting row number
        metadata: Dictionary of metadata key-value pairs

    Returns:
        Next available row number
    """
    row = start_row

    # Title
    ws.cell(row, 1, "Summary")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Metadata rows
    for key, value in metadata.items():
        ws.cell(row, 1, f"{key}:")
        ws.cell(row, 1).font = Font(bold=True)

        # Write value with appropriate type and formatting
        value_cell = ws.cell(row, 2)
        if isinstance(value, (list, tuple)):
            # Lists/tuples as comma-separated strings
            value_cell.value = ", ".join(str(v) for v in value)
        elif isinstance(value, int):
            # Integers as numbers with thousands separator
            value_cell.value = value
            value_cell.number_format = "#,##0"
        elif isinstance(value, float):
            # Floats as numbers with scientific notation
            value_cell.value = value
            value_cell.number_format = "0.00E+00"
        else:
            # Everything else as string
            value_cell.value = str(value)

        row += 1

    return row + 1  # Add spacing


def write_table_to_worksheet(
    ws: Worksheet,
    data: Dict[str, List],
    headers: List[str],
    title: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
    column_formats: Optional[Dict[str, str]] = None,
) -> None:
    """Write tabular data to worksheet with headers.

    Args:
        ws: Worksheet to write to
        data: Dictionary mapping header names to column data
        headers: List of column headers (order matters)
        title: Optional title for worksheet
        metadata: Optional summary data to append at bottom
        column_formats: Optional dictionary mapping column names to Excel number format codes
                       (e.g., {"Min BER": "0.00E+00"} for scientific notation)
    """
    row = 1

    # Write title if provided
    if title:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Write data rows
    num_rows = len(next(iter(data.values()))) if data else 0
    for data_row_idx in range(num_rows):
        for col_idx, header in enumerate(headers, start=1):
            value = data[header][data_row_idx]
            cell = ws.cell(row, col_idx, value)

            # Apply column-specific number formatting if specified
            # Only apply to numeric values (not empty strings or None)
            if column_formats and header in column_formats:
                if isinstance(value, (int, float)) and value != "":
                    cell.number_format = column_formats[header]
        row += 1

    # Adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 15

    # Write metadata section if provided
    if metadata:
        row += 1  # Spacing
        write_metadata_section(ws, row, metadata)


def write_heatmap_to_worksheet(
    ws: Worksheet,
    lane_data: Dict[str, Union[float, int]],
    color_scheme: ColorScheme,
    is_ber_metric: bool,
    title: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    """Write heatmap data with cell background colors.

    Args:
        ws: Worksheet to write to
        lane_data: Dictionary mapping lane_id to value
        color_scheme: Color scheme for heatmap
        is_ber_metric: True if BER metric, False if count metric
        title: Optional title for worksheet
        metadata: Optional summary data to append at bottom
    """
    row = 1

    # Write title if provided
    if title:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2

    # Group data by (bus_id, eth_id)
    grouped_data: Dict[Tuple[str, str], Dict[int, Union[float, int]]] = {}

    for lane_id, value in lane_data.items():
        # Parse lane_id: "hostname/bus_id/eth_id/lane_N" or "bus_id/eth_id/lane_N"
        parts = lane_id.split("/")
        if len(parts) == 4:
            # With hostname
            _, bus_id, eth_id, lane_str = parts
        elif len(parts) == 3:
            # Without hostname
            bus_id, eth_id, lane_str = parts
        else:
            logger.warning(f"Unexpected lane_id format: {lane_id}")
            continue

        # Extract lane number from "lane_N" or "laneN" format
        if "_" in lane_str:
            lane_num = int(lane_str.split("_")[1])
        else:
            # Handle "laneN" format (no underscore)
            lane_num = int(lane_str.replace("lane", ""))

        key = (bus_id, eth_id)
        if key not in grouped_data:
            grouped_data[key] = {}
        grouped_data[key][lane_num] = value

    # Write heatmap table
    # Header row: bus_id, eth_id, Lane 0, Lane 1, ..., Lane 7
    ws.cell(row, 1, "bus_id")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    ws.cell(row, 2, "eth_id")
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for lane_num in range(8):
        cell = ws.cell(row, lane_num + 3, f"Lane {lane_num}")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Data rows
    for (bus_id, eth_id), lanes in sorted(grouped_data.items()):
        # Port identifiers in separate columns
        ws.cell(row, 1, bus_id)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, eth_id)
        ws.cell(row, 2).font = Font(bold=True)

        # Lane values with colors
        for lane_num in range(8):
            value = lanes.get(lane_num)
            cell = ws.cell(row, lane_num + 3)  # +3 because of bus_id and eth_id columns

            if value is not None:
                # Store numeric value
                if is_ber_metric:
                    cell.value = value  # Store as number, not string
                    cell.number_format = "0.00E+00"  # Apply scientific notation formatting
                else:
                    cell.value = int(value)

                # Apply color (background only - use black text for all values)
                color = get_excel_color_for_value(value, color_scheme, is_ber_metric)
                apply_cell_background_color(cell, color)

                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")

        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 12  # bus_id
    ws.column_dimensions["B"].width = 12  # eth_id
    for col_idx in range(3, 11):  # Lane 0 through Lane 7
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 12

    # Write color legend
    row += 1
    ws.cell(row, 1, "Color Legend:")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1

    # Generate legend from thresholds
    sorted_thresholds = sorted(color_scheme.thresholds)
    for i, (threshold, color) in enumerate(sorted_thresholds):
        if is_ber_metric:
            threshold_text = f"<= {threshold:.2e}"
        else:
            threshold_text = f"<= {int(threshold)}" if threshold > 0 else "0"

        ws.cell(row, 1, threshold_text)
        color_hex = map_terminal_color_to_excel(color)
        apply_cell_background_color(ws.cell(row, 2, ""), color_hex)
        row += 1

    # Add default color for values exceeding all thresholds
    # Only show if default color is different from last threshold color
    last_threshold_color = sorted_thresholds[-1][1]
    if color_scheme.default_color != last_threshold_color:
        if is_ber_metric:
            threshold_text = f"> {sorted_thresholds[-1][0]:.2e}"
        else:
            threshold_text = f"> {int(sorted_thresholds[-1][0])}"
        ws.cell(row, 1, threshold_text)
        color_hex = map_terminal_color_to_excel(color_scheme.default_color)
        apply_cell_background_color(ws.cell(row, 2, ""), color_hex)
        row += 1

    # Write metadata section if provided
    if metadata:
        row += 1  # Spacing
        write_metadata_section(ws, row, metadata)


def write_histogram_to_worksheet(
    ws: Worksheet,
    histogram: Union[BERHistogram, List[BERHistogram]],
    title: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    """Write histogram as Excel column chart with bin legend.

    Args:
        ws: Worksheet to write to
        histogram: BERHistogram or list of histograms
        title: Optional title for worksheet
        metadata: Optional summary data to append at bottom
    """
    row = 1

    # Write title if provided
    if title:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2

    # Handle single histogram or list
    histograms = [histogram] if isinstance(histogram, BERHistogram) else histogram

    # Process each histogram
    for hist in histograms:
        # Write histogram title
        ws.cell(row, 1, f"BER Distribution - {hist.lane_id}")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1

        # Write data table
        ws.cell(row, 1, "BER Range")
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, "Count")
        ws.cell(row, 2).font = Font(bold=True)
        row += 1

        start_row = row
        for bin_label, count in hist.bins:
            ws.cell(row, 1, bin_label)
            ws.cell(row, 2, count)
            row += 1

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.title = f"BER Distribution - {hist.lane_id}"
        chart.x_axis.title = "BER Range"
        chart.y_axis.title = "Count"

        # Set chart dimensions (width x height in EMUs; 15 chars ≈ 1 inch, 20 rows ≈ 1 inch)
        chart.width = 18  # Width in inches
        chart.height = 10  # Height in inches

        # Style settings
        chart.style = 10  # Use a clean, professional style
        chart.legend = None  # Remove legend (not needed for single series)

        # Remove horizontal grid lines for cleaner appearance
        chart.y_axis.majorGridlines = None

        # Add data
        data = Reference(ws, min_col=2, min_row=start_row - 1, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Format the data series (bars)
        series = chart.series[0]
        series.graphicalProperties.solidFill = "4472C4"  # Professional blue color

        # Format individual data points with gradient colors based on bin position
        from openpyxl.drawing.fill import GradientFillProperties, GradientStop, ColorChoice
        num_bins = len(hist.bins)
        for idx in range(num_bins):
            # Create gradient from green (good) to red (bad)
            ratio = idx / max(num_bins - 1, 1)
            if ratio < 0.5:
                # Green to yellow
                r = int(76 + (255 - 76) * (ratio * 2))
                g = int(175 + (215 - 175) * (ratio * 2))
                b = int(80 + (0 - 80) * (ratio * 2))
            else:
                # Yellow to red
                r = 255
                g = int(215 - (215 - 0) * ((ratio - 0.5) * 2))
                b = 0

            hex_color = f"{r:02X}{g:02X}{b:02X}"
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = hex_color
            series.dPt.append(pt)

        # Position chart
        chart_cell = f"D{start_row - 1}"
        ws.add_chart(chart, chart_cell)

        # Add color legend below data table
        legend_start_row = row
        ws.cell(legend_start_row, 1, "Color Legend:")
        ws.cell(legend_start_row, 1).font = Font(bold=True)
        legend_start_row += 1

        # Create gradient legend entries (Low BER = Green, High BER = Red)
        legend_entries = [
            ("Low BER (Best)", "4CAF50"),      # Green
            ("", "9DC34D"),                    # Yellow-green
            ("Medium BER", "FFD700"),          # Yellow
            ("", "FFA500"),                    # Orange
            ("High BER (Worst)", "DC143C"),    # Red
        ]

        for label, hex_color in legend_entries:
            if label:  # Only show text for first, middle, and last entries
                ws.cell(legend_start_row, 1, label)
            # Apply background color to legend cell
            color_cell = ws.cell(legend_start_row, 2, "")
            color_cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            legend_start_row += 1

        row = legend_start_row + 2  # Spacing between histograms

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    # Write metadata section if provided
    if metadata:
        row += 1  # Spacing
        write_metadata_section(ws, row, metadata)


def write_ber_plot_to_worksheet(
    ws: Worksheet,
    plot: Union[BERPlot, List[BERPlot]],
    title: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    """Write BER plot as Excel line chart with data table.

    Data points are equally spaced on x-axis regardless of actual time intervals.

    Args:
        ws: Worksheet to write to
        plot: BERPlot or list of plots
        title: Optional title for worksheet
        metadata: Optional summary data to append at bottom
    """
    row = 1

    # Write title if provided
    if title:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2

    # Handle single plot or list
    plots = [plot] if isinstance(plot, BERPlot) else plot

    # Process each plot
    for ber_plot in plots:
        # Write plot title
        ws.cell(row, 1, f"BER Over Time - {ber_plot.lane_id}")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1

        # Write data table (Sample #, Timestamp, BER Value)
        ws.cell(row, 1, "Sample #")
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, "Timestamp")
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 3, "BER Value")
        ws.cell(row, 3).font = Font(bold=True)
        row += 1

        start_row = row
        for idx, point in enumerate(ber_plot.data_points, start=1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, point.timestamp)
            cell = ws.cell(row, 3, point.ber_value)
            cell.number_format = "0.00E+00"  # Scientific notation
            row += 1

        # Create line chart
        chart = LineChart()
        chart.title = f"BER Over Time - {ber_plot.lane_id}"
        chart.x_axis.title = "Sample Number"
        chart.y_axis.title = "BER Value"

        # Set chart dimensions
        chart.width = 18  # Width in inches
        chart.height = 10  # Height in inches

        # Style settings
        chart.style = 10  # Clean, professional style
        chart.legend = None  # Remove legend (single series)

        # Add data
        # Use Sample # as categories (x-axis), BER Value as data (y-axis)
        data = Reference(ws, min_col=3, min_row=start_row - 1, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Format the line (make it visible and professional)
        series = chart.series[0]
        series.graphicalProperties.line.solidFill = "4472C4"  # Professional blue
        series.graphicalProperties.line.width = 20000  # Line width in EMUs (thicker line)
        series.smooth = True  # Smooth line

        # Add data markers
        series.marker.symbol = "circle"
        series.marker.size = 5

        # Position chart to the right of the data table
        chart_cell = f"E{start_row - 1}"
        ws.add_chart(chart, chart_cell)

        row += 2  # Spacing between plots

    # Adjust column widths
    ws.column_dimensions["A"].width = 12  # Sample #
    ws.column_dimensions["B"].width = 20  # Timestamp
    ws.column_dimensions["C"].width = 15  # BER Value

    # Write metadata section if provided
    if metadata:
        row += 1  # Spacing
        write_metadata_section(ws, row, metadata)
