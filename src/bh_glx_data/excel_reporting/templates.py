"""Excel template management and manipulation."""

import logging
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from bh_glx_data.core.exceptions import ExcelGenerationError, TemplateError

logger = logging.getLogger(__name__)

# Default sheet names in template
DEFAULT_SHEET_RAW_PRBS = "raw prbs data"
DEFAULT_SHEET_RAW_DATA = "raw data"
DEFAULT_SHEET_PRBS_SUMMARY = "PRBS Summary"
DEFAULT_SHEET_DATA_SUMMARY = "DATA Summary"


def load_template(template_path: Path) -> openpyxl.Workbook:
    """Load Excel template workbook.

    Args:
        template_path: Path to template file

    Returns:
        Loaded openpyxl Workbook

    Raises:
        TemplateError: If template cannot be loaded
    """
    if not template_path.exists():
        raise TemplateError(
            f"Template file not found: {template_path}",
            template_path=str(template_path),
        )

    try:
        workbook = openpyxl.load_workbook(template_path)
        logger.info(f"Loaded template from {template_path}")
        return workbook
    except Exception as e:
        raise TemplateError(
            f"Error loading template: {e}",
            template_path=str(template_path),
        )


def paste_data_to_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    data_df: pd.DataFrame,
) -> Tuple[int, int]:
    """Paste compiled data to specified sheet.

    Overwrites existing data in the sheet starting from row 1.

    Args:
        workbook: openpyxl Workbook object
        sheet_name: Name of the sheet to paste data into
        data_df: DataFrame to paste

    Returns:
        Tuple of (last_row, last_col) with 1-based indices

    Raises:
        ExcelGenerationError: If data pasting fails
    """
    if sheet_name not in workbook.sheetnames:
        raise ExcelGenerationError(
            f"Sheet '{sheet_name}' not found in workbook",
        )

    try:
        sheet = workbook[sheet_name]

        # Clear existing data if there is any
        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)

        # Write headers
        headers = list(data_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = str(header) if header is not None else ""

        # Write data rows
        for row_idx, (_, row_data) in enumerate(data_df.iterrows(), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # Handle NaN, None, and other special values
                if pd.isna(value):
                    cell.value = None
                elif value is None:
                    cell.value = None
                else:
                    # Convert to string for complex objects, otherwise keep as-is
                    try:
                        cell.value = value
                    except (TypeError, ValueError):
                        cell.value = str(value)

        last_row = len(data_df) + 1  # +1 for header row
        last_col = len(headers)

        logger.info(
            f"Pasted {len(data_df)} rows to sheet '{sheet_name}' "
            f"(range: A1:{get_column_letter(last_col)}{last_row})"
        )
        return last_row, last_col

    except Exception as e:
        logger.error(f"Error pasting data to sheet '{sheet_name}': {e}")
        raise ExcelGenerationError(
            f"Failed to paste data to sheet '{sheet_name}': {e}",
        )


def update_pivot_table_source(
    workbook: openpyxl.Workbook,
    pivot_sheet_name: str,
    data_sheet_name: str,
    data_range: str,
) -> bool:
    """Update pivot table data source and refresh.

    Args:
        workbook: openpyxl Workbook object
        pivot_sheet_name: Name of the sheet containing the pivot table
        data_sheet_name: Name of the sheet containing the data
        data_range: Excel range string (e.g., 'A1:Z1000')

    Returns:
        True if successful, False otherwise
    """
    if pivot_sheet_name not in workbook.sheetnames:
        logger.error(f"Pivot sheet '{pivot_sheet_name}' not found in workbook")
        return False

    try:
        sheet = workbook[pivot_sheet_name]

        # Access pivot tables through the sheet
        # In openpyxl, pivot tables are accessed via sheet._pivots (private attribute)
        if not hasattr(sheet, "_pivots") or not sheet._pivots:
            logger.warning(f"No pivot tables found in sheet '{pivot_sheet_name}'")
            return False

        # Update each pivot table's data source and refresh
        updated_count = 0
        for pivot_table in sheet._pivots:
            try:
                # Access the cache
                cache = pivot_table.cache
                if cache and hasattr(cache, "cacheSource"):
                    source = cache.cacheSource
                    if hasattr(source, "worksheetSource"):
                        ws_source = source.worksheetSource
                        # Update the range and sheet name
                        ws_source.ref = data_range
                        ws_source.sheet = data_sheet_name

                        # Refresh the pivot table cache
                        # Mark the cache as needing refresh by updating its refreshOnLoad flag
                        if hasattr(cache, "refreshOnLoad"):
                            cache.refreshOnLoad = True
                        # Also try to invalidate the cache
                        if hasattr(cache, "refresh"):
                            try:
                                cache.refresh()
                            except (AttributeError, TypeError):
                                pass  # Some cache objects may not have refresh method

                        updated_count += 1
                        logger.info(
                            f"Updated and refreshed pivot table data source to "
                            f"'{data_sheet_name}'!{data_range}"
                        )
            except Exception as e:
                logger.warning(f"Could not update one pivot table: {e}")
                continue

        if updated_count > 0:
            return True
        else:
            logger.warning(f"Could not update any pivot tables in sheet '{pivot_sheet_name}'")
            return False

    except Exception as e:
        logger.error(f"Error updating pivot table source in sheet '{pivot_sheet_name}': {e}")
        return False


def refresh_pivot_tables(workbook: openpyxl.Workbook, sheet_name: str) -> bool:
    """Refresh all pivot tables in the specified sheet.

    Args:
        workbook: openpyxl Workbook object
        sheet_name: Name of the sheet containing pivot tables

    Returns:
        True if successful, False otherwise
    """
    if sheet_name not in workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found, skipping pivot table refresh")
        return False

    try:
        sheet = workbook[sheet_name]

        if not hasattr(sheet, "_pivots") or not sheet._pivots:
            logger.debug(f"No pivot tables found in sheet '{sheet_name}'")
            return True  # Not an error, just no pivot tables

        refreshed_count = 0
        for pivot_table in sheet._pivots:
            try:
                cache = pivot_table.cache
                if cache:
                    # Set refreshOnLoad flag so Excel will refresh when opened
                    if hasattr(cache, "refreshOnLoad"):
                        cache.refreshOnLoad = True
                    refreshed_count += 1
            except Exception as e:
                logger.warning(f"Could not refresh one pivot table: {e}")
                continue

        if refreshed_count > 0:
            logger.info(
                f"Marked {refreshed_count} pivot table(s) for refresh in sheet '{sheet_name}'"
            )

        return True

    except Exception as e:
        logger.error(f"Error refreshing pivot tables in sheet '{sheet_name}': {e}")
        return False


def save_workbook(workbook: openpyxl.Workbook, output_path: Path) -> None:
    """Save workbook to file.

    Args:
        workbook: openpyxl Workbook object
        output_path: Path where to save the workbook

    Raises:
        ExcelGenerationError: If workbook cannot be saved
    """
    try:
        workbook.save(output_path)
        logger.info(f"Saved Excel file to {output_path}")
    except Exception as e:
        raise ExcelGenerationError(
            f"Failed to save Excel file: {e}",
            output_path=str(output_path),
        )
