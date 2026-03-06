"""Excel generation logic for test data summaries."""

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter

from bh_glx_data.core.exceptions import DataProcessingError, ExcelGenerationError
from bh_glx_data.excel_reporting.templates import (
    load_template,
    paste_data_to_sheet,
    refresh_pivot_tables,
    save_workbook,
    update_pivot_table_source,
)

logger = logging.getLogger(__name__)

# Test type constants
TEST_TYPE_PRBS = "TestType.SERDES_PRBS"
TEST_TYPE_DATA = "TestType.SIMPLE_PACKET"

# Default sheet names
SHEET_RAW_PRBS = "raw prbs data"
SHEET_RAW_DATA = "raw data"
SHEET_PRBS_SUMMARY = "PRBS Summary"
SHEET_DATA_SUMMARY = "DATA Summary"


def extract_firmware_version(csv_path: Path) -> str:
    """Extract firmware version from CSV filename.

    Looks for pattern like 'erisc_v1_7_103' or 'v1_7_103' in filename.

    Args:
        csv_path: Path to CSV file

    Returns:
        Firmware version string (e.g., 'erisc_v1_7_103') or 'unknown' if not found
    """
    filename = csv_path.name

    # Try to match erisc_vX_Y_Z pattern first
    pattern1 = r"erisc_v\d+_\d+_\d+"
    match = re.search(pattern1, filename)
    if match:
        return match.group(0)

    # Try to match vX_Y_Z pattern
    pattern2 = r"v\d+_\d+_\d+"
    match = re.search(pattern2, filename)
    if match:
        return match.group(0)

    logger.warning(f"Could not extract firmware version from filename: {filename}, using 'unknown'")
    return "unknown"


def identify_test_type(csv_path: Path) -> Optional[str]:
    """Determine if CSV is PRBS or Data test by reading the test_type column.

    Args:
        csv_path: Path to CSV file

    Returns:
        'PRBS' or 'DATA' or None if cannot determine
    """
    try:
        # Read just the first row to check test_type
        df = pd.read_csv(csv_path, nrows=1)

        if "test_type" not in df.columns:
            logger.warning(f"CSV file {csv_path.name} does not have 'test_type' column")
            # Fallback to filename check
            if "prbs_test" in csv_path.name.lower():
                return "PRBS"
            elif "data_test" in csv_path.name.lower():
                return "DATA"
            return None

        test_type_value = df["test_type"].iloc[0]

        if test_type_value == TEST_TYPE_PRBS:
            return "PRBS"
        elif test_type_value == TEST_TYPE_DATA:
            return "DATA"
        else:
            logger.warning(f"Unknown test_type '{test_type_value}' in {csv_path.name}")
            # Fallback to filename check
            if "prbs_test" in csv_path.name.lower():
                return "PRBS"
            elif "data_test" in csv_path.name.lower():
                return "DATA"
            return None

    except Exception as e:
        logger.error(f"Error reading CSV file {csv_path.name}: {e}")
        # Fallback to filename check
        if "prbs_test" in csv_path.name.lower():
            return "PRBS"
        elif "data_test" in csv_path.name.lower():
            return "DATA"
        return None


def extract_system_hostname(csv_path: Path) -> Optional[str]:
    """Extract system hostname from CSV host column.

    Args:
        csv_path: Path to CSV file

    Returns:
        System hostname (e.g., 'bh-glx-b02u02') or None if not found
    """
    try:
        # Read just the first row to get hostname
        df = pd.read_csv(csv_path, nrows=1)

        if "host" not in df.columns:
            logger.warning(f"CSV file {csv_path.name} does not have 'host' column")
            return None

        hostname = df["host"].iloc[0]
        if pd.isna(hostname) or hostname == "":
            logger.warning(f"Empty hostname in {csv_path.name}")
            return None

        return str(hostname).strip()

    except Exception as e:
        logger.error(f"Error extracting hostname from {csv_path.name}: {e}")
        return None


def scan_csv_files(data_dir: Path) -> List[Path]:
    """Scan directory and return list of CSV files.

    Args:
        data_dir: Directory to scan

    Returns:
        List of Path objects for CSV files

    Raises:
        DataProcessingError: If directory doesn't exist
    """
    if not data_dir.exists():
        raise DataProcessingError(f"Data directory {data_dir} does not exist")

    csv_files = list(data_dir.glob("*.csv"))
    logger.info(f"Found {len(csv_files)} CSV files in {data_dir}")
    return csv_files


def group_csvs_by_system(csv_files: List[Path]) -> Dict[Tuple[str, str], Dict[str, List[Path]]]:
    """Group CSV files by system hostname AND firmware version.

    Args:
        csv_files: List of Path objects for CSV files

    Returns:
        Dictionary with keys (hostname, firmware_version) and values as
        dicts with 'PRBS' and 'DATA' lists of CSV paths
    """
    grouped: Dict[Tuple[str, str], Dict[str, List[Path]]] = defaultdict(
        lambda: {"PRBS": [], "DATA": []}
    )

    for csv_path in csv_files:
        hostname = extract_system_hostname(csv_path)
        firmware_version = extract_firmware_version(csv_path)
        test_type = identify_test_type(csv_path)

        if not hostname:
            logger.warning(f"Skipping {csv_path.name}: could not extract hostname")
            continue

        if not test_type:
            logger.warning(f"Skipping {csv_path.name}: could not identify test type")
            continue

        key = (hostname, firmware_version)
        grouped[key][test_type].append(csv_path)
        logger.debug(
            f"Grouped {csv_path.name}: hostname={hostname}, "
            f"firmware={firmware_version}, type={test_type}"
        )

    return grouped


def compile_test_data(csv_files: List[Path], test_type: str) -> Optional[pd.DataFrame]:
    """Combine CSV data for a test type by reading and concatenating all CSV files.

    Args:
        csv_files: List of Path objects for CSV files of the same test type
        test_type: 'PRBS' or 'DATA'

    Returns:
        Combined DataFrame with all rows from all CSV files, or None if error
    """
    if not csv_files:
        logger.warning(f"No CSV files provided for {test_type} test type")
        return None

    dataframes = []

    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path)
            if df.empty:
                logger.warning(f"CSV file {csv_path.name} is empty")
                continue
            dataframes.append(df)
            logger.debug(f"Loaded {len(df)} rows from {csv_path.name}")
        except Exception as e:
            logger.error(f"Error reading CSV file {csv_path.name}: {e}")
            continue

    if not dataframes:
        logger.warning(f"No valid data found for {test_type} test type")
        return None

    # Concatenate all dataframes
    combined_df = pd.concat(dataframes, ignore_index=True)
    logger.info(
        f"Compiled {len(combined_df)} total rows for {test_type} test type "
        f"from {len(dataframes)} files"
    )

    return combined_df


def generate_excel_summary(
    hostname: str,
    firmware_version: str,
    prbs_data: Optional[pd.DataFrame],
    data_test_data: Optional[pd.DataFrame],
    template_path: Path,
    output_dir: Path,
) -> Path:
    """Generate Excel summary file for a system+firmware combination.

    Args:
        hostname: System hostname (e.g., 'bh-glx-b02u02')
        firmware_version: Firmware version (e.g., 'erisc_v1_7_103')
        prbs_data: PRBS test data or None
        data_test_data: Data test data or None
        template_path: Path to Excel template file
        output_dir: Directory where to save the output file

    Returns:
        Path to the generated Excel file

    Raises:
        ExcelGenerationError: If Excel generation fails
    """
    # Load template
    workbook = load_template(template_path)

    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate output filename
    output_filename = f"{hostname}_{firmware_version}.xlsx"
    output_path = output_dir / output_filename

    # Paste PRBS data if available
    if prbs_data is not None and not prbs_data.empty:
        last_row, last_col = paste_data_to_sheet(workbook, SHEET_RAW_PRBS, prbs_data)
        prbs_range = f"A1:{get_column_letter(last_col)}{last_row}"

        # Update PRBS Summary pivot table if sheet exists
        if SHEET_PRBS_SUMMARY in workbook.sheetnames:
            update_pivot_table_source(workbook, SHEET_PRBS_SUMMARY, SHEET_RAW_PRBS, prbs_range)
            refresh_pivot_tables(workbook, SHEET_PRBS_SUMMARY)
        else:
            logger.warning(f"Sheet '{SHEET_PRBS_SUMMARY}' not found, skipping pivot table update")
    else:
        logger.info(f"No PRBS data for {hostname} {firmware_version}")

    # Paste Data test data if available
    if data_test_data is not None and not data_test_data.empty:
        last_row, last_col = paste_data_to_sheet(workbook, SHEET_RAW_DATA, data_test_data)
        data_range = f"A1:{get_column_letter(last_col)}{last_row}"

        # Update DATA Summary pivot table if sheet exists
        if SHEET_DATA_SUMMARY in workbook.sheetnames:
            update_pivot_table_source(workbook, SHEET_DATA_SUMMARY, SHEET_RAW_DATA, data_range)
            refresh_pivot_tables(workbook, SHEET_DATA_SUMMARY)
        else:
            logger.warning(f"Sheet '{SHEET_DATA_SUMMARY}' not found, skipping pivot table update")
    else:
        logger.info(f"No Data test data for {hostname} {firmware_version}")

    # Save workbook
    save_workbook(workbook, output_path)

    return output_path


def process_all_systems(
    data_dir: Path,
    template_path: Path,
    output_dir: Path,
    system_filter: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Process all systems and generate Excel summaries.

    Args:
        data_dir: Directory containing CSV files
        template_path: Path to Excel template
        output_dir: Directory for output Excel files
        system_filter: Optional list of hostnames to filter by

    Returns:
        Dictionary with processing summary:
        {
            'total_combinations': int,
            'success_count': int,
            'error_count': int,
            'generated_files': List[Path],
            'errors': List[str]
        }
    """
    # Scan CSV files
    csv_files = scan_csv_files(data_dir)

    if not csv_files:
        raise DataProcessingError("No CSV files found in data directory")

    # Group CSV files by system and firmware
    grouped = group_csvs_by_system(csv_files)

    if not grouped:
        raise DataProcessingError("No valid CSV files could be grouped")

    # Filter by systems if specified
    if system_filter:
        systems_to_process = set(system_filter)
        filtered_grouped = {
            (hostname, firmware_version): file_groups
            for (hostname, firmware_version), file_groups in grouped.items()
            if hostname in systems_to_process
        }

        if not filtered_grouped:
            raise DataProcessingError(
                f"No data found for specified systems: {', '.join(system_filter)}"
            )

        # Check if any requested systems were not found
        found_systems = {hostname for hostname, _ in filtered_grouped.keys()}
        missing_systems = systems_to_process - found_systems
        if missing_systems:
            logger.warning(f"No data found for systems: {', '.join(missing_systems)}")

        grouped = filtered_grouped
        logger.info(f"Filtering to {len(grouped)} system+firmware combinations")

    # Process each system+firmware combination
    success_count = 0
    error_count = 0
    generated_files = []
    errors = []

    for (hostname, firmware_version), file_groups in grouped.items():
        logger.info(f"\nProcessing {hostname} with firmware {firmware_version}")

        try:
            # Compile PRBS data
            prbs_data = None
            if file_groups["PRBS"]:
                prbs_data = compile_test_data(file_groups["PRBS"], "PRBS")

            # Compile Data test data
            data_test_data = None
            if file_groups["DATA"]:
                data_test_data = compile_test_data(file_groups["DATA"], "DATA")

            # Generate Excel file
            if prbs_data is None and data_test_data is None:
                logger.warning(f"No data available for {hostname} {firmware_version}, skipping")
                error_count += 1
                errors.append(f"No data for {hostname} {firmware_version}")
                continue

            output_path = generate_excel_summary(
                hostname,
                firmware_version,
                prbs_data,
                data_test_data,
                template_path,
                output_dir,
            )

            generated_files.append(output_path)
            success_count += 1

        except Exception as e:
            logger.error(f"Error processing {hostname} {firmware_version}: {e}")
            error_count += 1
            errors.append(f"{hostname} {firmware_version}: {e}")

    return {
        "total_combinations": len(grouped),
        "success_count": success_count,
        "error_count": error_count,
        "generated_files": generated_files,
        "errors": errors,
    }
