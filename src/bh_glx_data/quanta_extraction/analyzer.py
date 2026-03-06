"""Excel analysis logic for Quanta QC3 test results."""

import logging
from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from tqdm import tqdm

    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

from bh_glx_data.core.exceptions import DataProcessingError
from bh_glx_data.core.models import QC3TestResult

logger = logging.getLogger(__name__)


def _get_cell_value(cell) -> str:
    """Safely get cell value, returning empty string if None."""
    if cell is None:
        return ""
    value = cell.value
    return "" if value is None else str(value).strip()


def _get_numeric_value(cell) -> int:
    """Get numeric value from cell as integer, returning 0 if not numeric."""
    if cell is None or cell.value is None:
        return 0
    try:
        return int(float(cell.value))
    except (ValueError, TypeError):
        return 0


def analyze_excel_failures(excel_path: Path) -> List[QC3TestResult]:
    """Read Excel file and return rows with non-zero Failure Count.

    Reads QC3 test Excel file and extracts:
    - SN (Serial Number) from Column C
    - ASIC from Column J
    - PORT from Column K
    - TYPE from Column M
    - Failure Count from Column N

    Args:
        excel_path: Path to the Excel file

    Returns:
        List of QC3TestResult objects for rows with failures

    Raises:
        DataProcessingError: If Excel file cannot be read
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise DataProcessingError(f"Excel file not found: {excel_path}")

    if not (HAS_OPENPYXL or HAS_PANDAS):
        raise DataProcessingError(
            "Neither openpyxl nor pandas is available. Please install openpyxl: pip install openpyxl"
        )

    try:
        failures = []

        if HAS_OPENPYXL:
            # Use openpyxl method
            logger.info(f"Reading Excel file with openpyxl: {excel_path.name}")
            workbook = load_workbook(excel_path, read_only=True, data_only=True)

            # Get the first (active) sheet
            sheet = workbook.active

            # Column indices (Excel is 1-indexed)
            # Column C = 3, Column J = 10, Column K = 11, Column M = 13, Column N = 14
            SN_COL = 3
            ASIC_COL = 10
            PORT_COL = 11
            TYPE_COL = 13
            FCOUNT_COL = 14

            # Read data starting from row 2 (assuming row 1 is header)
            max_row = sheet.max_row
            total_rows = max_row - 1  # Excluding header row

            # Create progress bar
            row_iterator = range(2, max_row + 1)
            if HAS_TQDM:
                row_iterator = tqdm(
                    row_iterator,
                    desc="Processing rows",
                    unit="row",
                    total=total_rows,
                )

            for row_num in row_iterator:
                # Get Failure Count value (Column N)
                fcount_cell = sheet.cell(row=row_num, column=FCOUNT_COL)
                fcount = _get_numeric_value(fcount_cell)

                # Only process rows with non-zero Failure Count
                if fcount != 0:
                    sn = _get_cell_value(sheet.cell(row=row_num, column=SN_COL))
                    asic = _get_cell_value(sheet.cell(row=row_num, column=ASIC_COL))
                    port = _get_cell_value(sheet.cell(row=row_num, column=PORT_COL))
                    type_val = _get_cell_value(sheet.cell(row=row_num, column=TYPE_COL))

                    failures.append(
                        QC3TestResult(
                            serial_number=sn,
                            asic=asic,
                            port=port,
                            failure_count=fcount,
                            test_type=type_val,
                        )
                    )

            workbook.close()

        elif HAS_PANDAS:
            # Fallback to pandas method
            logger.info(f"Reading Excel file with pandas: {excel_path.name}")
            if HAS_TQDM:
                logger.info("Reading Excel file...")

            try:
                # Try with openpyxl engine first
                df = pd.read_excel(excel_path, engine="openpyxl", header=0)
            except (ImportError, ValueError):
                # If openpyxl not available, try default engine
                df = pd.read_excel(excel_path, header=0)

            # Column indices (0-indexed in pandas)
            # Column C = index 2, Column J = index 9, Column K = index 10,
            # Column M = index 12, Column N = index 13
            columns = df.columns.tolist()

            if len(columns) < 14:
                raise DataProcessingError(f"Expected at least 14 columns, found {len(columns)}")

            # Get column values by index
            sn_col = columns[2]  # Column C
            asic_col = columns[9]  # Column J
            port_col = columns[10]  # Column K
            type_col = columns[12]  # Column M
            fcount_col = columns[13]  # Column N

            # Filter rows where Failure Count is non-zero
            if HAS_TQDM:
                logger.info("Filtering rows...")

            df[fcount_col] = pd.to_numeric(df[fcount_col], errors="coerce").fillna(0)
            failures_df = df[df[fcount_col] != 0]

            # Convert to list of QC3TestResult objects
            row_iterator = failures_df.iterrows()
            if HAS_TQDM:
                row_iterator = tqdm(
                    row_iterator,
                    desc="Processing failures",
                    unit="row",
                    total=len(failures_df),
                )

            for idx, row in row_iterator:
                failures.append(
                    QC3TestResult(
                        serial_number=str(row[sn_col]) if pd.notna(row[sn_col]) else "",
                        asic=str(row[asic_col]) if pd.notna(row[asic_col]) else "",
                        port=str(row[port_col]) if pd.notna(row[port_col]) else "",
                        failure_count=int(row[fcount_col]),
                        test_type=str(row[type_col]) if pd.notna(row[type_col]) else "",
                    )
                )

        logger.info(f"Found {len(failures)} failure(s) in Excel file")
        return failures

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise DataProcessingError(f"Failed to analyze Excel file: {e}")


def print_failures_summary(failures: List[QC3TestResult]) -> None:
    """Print a formatted summary of failures.

    Args:
        failures: List of QC3TestResult objects
    """
    if not failures:
        print("No failures found (all Failure Count values are zero).")
        return

    # Print header
    print(f"{'SN':<20} {'ASIC':<15} {'PORT':<10} {'Failure Count':<15} {'TYPE':<20}")
    print("-" * 85)

    # Print each failure row
    for failure in failures:
        print(
            f"{failure.serial_number:<20} {failure.asic:<15} {failure.port:<10} "
            f"{failure.failure_count:<15} {failure.test_type:<20}"
        )

    print(f"\nTotal failures found: {len(failures)}")


def extract_failed_serial_numbers(failures: List[QC3TestResult]) -> List[str]:
    """Extract unique serial numbers from failures.

    Args:
        failures: List of QC3TestResult objects

    Returns:
        List of unique serial numbers with failures
    """
    serial_numbers = set()
    for failure in failures:
        if failure.serial_number:
            serial_numbers.add(failure.serial_number)

    return sorted(list(serial_numbers))
