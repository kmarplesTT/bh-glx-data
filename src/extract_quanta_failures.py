#!/usr/bin/env python3
"""
Extract test data for failed systems from Quanta

This tool:
1. Unzips a Quanta test package
2. Analyzes the Excel test results to find systems with failures
3. Locates directories containing the failed system serial numbers
4. Extracts data_test_*.csv and prbs_test_*.csv files from nested archives
5. Stores the extracted CSV files in the quanta/ directory

File Organization:
- QC3_*.zip contains:
  - QC3_*_test.xlsx (main results spreadsheet)
  - 0130/{SN1}_{SN2}_{SN3}_{SN4}/
    - QC3_UBB_*.tar.gz contains:
      - tt_funtest_ubb_*/
        - ft_eth_stress_*.tar.gz contains:
          - ft_eth_stress/
            - data_test_*.csv
            - prbs_test_*.csv
"""

import os
import sys
import re
import zipfile
import tarfile
import tempfile
import shutil
from pathlib import Path
from typing import List, Set, Dict, Tuple
import subprocess

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


def get_cell_value(cell):
    """Safely get cell value, returning empty string if None."""
    if cell is None:
        return ""
    value = cell.value
    return "" if value is None else str(value).strip()


def get_numeric_value(cell):
    """Get numeric value from cell as integer, returning 0 if not numeric."""
    if cell is None or cell.value is None:
        return 0
    try:
        return int(float(cell.value))
    except (ValueError, TypeError):
        return 0


def get_failed_serial_numbers(excel_path: Path) -> Set[str]:
    """
    Extract serial numbers of systems with failures from Excel file.

    Args:
        excel_path: Path to the QC3 Excel test results file

    Returns:
        Set of serial numbers that have non-zero failure counts
    """
    if not HAS_OPENPYXL:
        print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    print(f"Analyzing failures in {excel_path.name}...")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    # Column indices (Excel is 1-indexed)
    # Column C = 3 (SN), Column N = 14 (Failure Count)
    SN_COL = 3
    FCOUNT_COL = 14

    failed_sns = set()

    # Read data starting from row 2 (assuming row 1 is header)
    max_row = sheet.max_row
    row_iterator = range(2, max_row + 1)

    if HAS_TQDM:
        row_iterator = tqdm(row_iterator, desc="Scanning for failures", unit="row")

    for row_num in row_iterator:
        fcount_cell = sheet.cell(row=row_num, column=FCOUNT_COL)
        fcount = get_numeric_value(fcount_cell)

        if fcount != 0:
            sn = get_cell_value(sheet.cell(row=row_num, column=SN_COL))
            if sn:
                failed_sns.add(sn)

    workbook.close()

    return failed_sns


def find_matching_directories(extract_dir: Path, serial_numbers: Set[str]) -> List[Path]:
    """
    Find directories that contain any of the failed serial numbers in their name.

    Args:
        extract_dir: Root directory to search
        serial_numbers: Set of serial numbers to look for

    Returns:
        List of directory paths that match
    """
    matching_dirs = []

    # Look for date-based subdirectories (e.g., "0130")
    for date_dir in extract_dir.iterdir():
        if not date_dir.is_dir():
            continue

        # Check each subdirectory with combined serial numbers
        for sn_dir in date_dir.iterdir():
            if not sn_dir.is_dir():
                continue

            # Check if directory name contains any of the failed serial numbers
            dir_name = sn_dir.name
            for sn in serial_numbers:
                if sn in dir_name:
                    matching_dirs.append(sn_dir)
                    break

    return matching_dirs


def extract_csv_files(sn_dir: Path, output_dir: Path) -> int:
    """
    Extract data_test_*.csv and prbs_test_*.csv files from a system directory.

    Handles nested tar.gz structure:
    - Finds QC3_UBB_*.tar.gz files
    - Extracts tt_funtest_ubb_*/ft_eth_stress_*.tar.gz
    - Extracts ft_eth_stress/data_test_*.csv and prbs_test_*.csv

    Args:
        sn_dir: Directory containing the serial number test data
        output_dir: Destination directory for extracted CSV files

    Returns:
        Number of CSV files extracted
    """
    csv_count = 0

    # Find QC3_UBB_*.tar.gz files
    ubb_archives = list(sn_dir.glob("QC3_UBB_*.tar.gz"))

    if not ubb_archives:
        print(f"  Warning: No QC3_UBB_*.tar.gz found in {sn_dir.name}")
        return 0

    for ubb_archive in ubb_archives:
        print(f"  Processing {ubb_archive.name}...")

        # Create temporary directory for extraction
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            try:
                # Extract the UBB archive
                with tarfile.open(ubb_archive, 'r:gz') as tar:
                    # Find tt_funtest_ubb_* directories
                    funtest_members = [m for m in tar.getmembers()
                                      if 'tt_funtest_ubb_' in m.name and
                                      'ft_eth_stress_' in m.name and
                                      m.name.endswith('.tar.gz')]

                    if not funtest_members:
                        print(f"    Warning: No ft_eth_stress_*.tar.gz found")
                        continue

                    # Extract the ft_eth_stress archive
                    for member in funtest_members:
                        tar.extract(member, temp_path)
                        eth_stress_archive = temp_path / member.name

                        # Extract CSV files from ft_eth_stress archive
                        with tarfile.open(eth_stress_archive, 'r:gz') as eth_tar:
                            csv_members = [m for m in eth_tar.getmembers()
                                          if m.name.startswith('ft_eth_stress/') and
                                          (m.name.endswith('data_test_' + m.name.split('data_test_')[-1]) or
                                           m.name.endswith('prbs_test_' + m.name.split('prbs_test_')[-1])) and
                                          m.name.endswith('.csv')]

                            for csv_member in csv_members:
                                # Extract to output directory with a unique name
                                csv_filename = Path(csv_member.name).name
                                # Prefix with the system directory name to avoid collisions
                                output_filename = f"{sn_dir.name}_{csv_filename}"
                                output_path = output_dir / output_filename

                                # Extract the file
                                eth_tar.extract(csv_member, temp_path)
                                extracted_file = temp_path / csv_member.name

                                # Copy to output directory
                                shutil.copy2(extracted_file, output_path)
                                print(f"    Extracted: {output_filename}")
                                csv_count += 1

            except (tarfile.TarError, OSError) as e:
                print(f"    Error extracting from {ubb_archive.name}: {e}", file=sys.stderr)
                continue

    return csv_count


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description='Extract test data for failed systems from Quanta QC3 test packages',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Extract failures from a specific test package
  %(prog)s QC3_UBB_20260128_build.zip

  # Specify custom output directory
  %(prog)s QC3_UBB_20260128_build.zip --output-dir my_failures/

The tool will:
  1. Unzip the test package
  2. Find the Excel results file and identify failed systems
  3. Extract CSV test data for those systems to quanta/ directory
"""
    )

    parser.add_argument(
        'zip_file',
        type=str,
        help='Path to the QC3 test package zip file (e.g., QC3_UBB_20260128_build.zip)'
    )

    parser.add_argument(
        '-o', '--output-dir',
        type=str,
        default='quanta',
        help='Output directory for extracted CSV files (default: quanta/)'
    )

    args = parser.parse_args()

    # Validate input file
    zip_path = Path(args.zip_file)
    if not zip_path.exists():
        print(f"Error: File not found: {zip_path}", file=sys.stderr)
        sys.exit(1)

    if not zip_path.suffix == '.zip':
        print(f"Error: Expected a .zip file, got: {zip_path.suffix}", file=sys.stderr)
        sys.exit(1)

    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {output_dir.absolute()}")

    # Create temporary directory for extraction
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)

        print(f"\nExtracting {zip_path.name}...")
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_path)
        except zipfile.BadZipFile:
            print(f"Error: Invalid zip file: {zip_path}", file=sys.stderr)
            sys.exit(1)

        # Find the root directory (should be something like QC3_UBB_20260128_build)
        extract_dirs = [d for d in temp_path.iterdir() if d.is_dir()]
        if not extract_dirs:
            print("Error: No directories found in zip file", file=sys.stderr)
            sys.exit(1)

        extract_dir = extract_dirs[0]

        # Find the Excel file
        excel_files = list(extract_dir.glob("*.xlsx"))
        if not excel_files:
            print(f"Error: No Excel file found in {extract_dir}", file=sys.stderr)
            sys.exit(1)

        excel_path = excel_files[0]

        # Get failed serial numbers
        failed_sns = get_failed_serial_numbers(excel_path)

        if not failed_sns:
            print("\nNo failures found in test results.")
            return

        print(f"\nFound {len(failed_sns)} system(s) with failures:")
        for sn in sorted(failed_sns):
            print(f"  - {sn}")

        # Find directories containing failed systems
        print("\nSearching for test data directories...")
        matching_dirs = find_matching_directories(extract_dir, failed_sns)

        if not matching_dirs:
            print("Warning: No directories found matching failed serial numbers", file=sys.stderr)
            return

        print(f"Found {len(matching_dirs)} matching director{'y' if len(matching_dirs) == 1 else 'ies'}")

        # Extract CSV files from each matching directory
        total_csv_count = 0
        print("\nExtracting CSV files...")
        for sn_dir in matching_dirs:
            print(f"\nProcessing {sn_dir.name}...")
            csv_count = extract_csv_files(sn_dir, output_dir)
            total_csv_count += csv_count

        print(f"\n{'='*80}")
        print(f"Extraction complete!")
        print(f"Total CSV files extracted: {total_csv_count}")
        print(f"Output directory: {output_dir.absolute()}")
        print(f"{'='*80}")


if __name__ == "__main__":
    main()
