#!/usr/bin/env python3
"""
Analyze failures from QC3 test Excel file.

Reads QC3_S7TK_0128_build_test.xlsx and prints out rows where Failure Count (Column N) is non-zero,
showing SN (Column C), ASIC (Column J), PORT (Column K), Failure Count (Column N), and TYPE (Column M).
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    try:
        import pandas as pd
        HAS_PANDAS = True
    except ImportError:
        HAS_PANDAS = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


def get_cell_value(cell):
    """Safely get cell value, returning empty string if None."""
    if cell is None:
        return ""
    value = cell.value
    return "" if value is None else str(value).strip()


def get_numeric_value(cell):
    """Get numeric value from cell as integer, returning 0 if not numeric."""
    if cell is None or cell.value is None:
        return 0
    try:
        return int(float(cell.value))
    except (ValueError, TypeError):
        return 0


def analyze_failures(excel_path):
    """
    Read Excel file and print rows with non-zero Failure Count.

    Args:
        excel_path (str or Path): Path to the Excel file
    """
    # Convert to Path object
    excel_path = Path(excel_path)

    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    try:
        if HAS_OPENPYXL:
            # Use openpyxl method
            workbook = load_workbook(excel_path, read_only=True, data_only=True)

            # Get the first (active) sheet
            sheet = workbook.active

            # Column indices (Excel is 1-indexed)
            # Column C = 3, Column J = 10, Column K = 11, Column M = 13, Column N = 14
            SN_COL = 3
            ASIC_COL = 10
            PORT_COL = 11
            TYPE_COL = 13
            FCOUNT_COL = 14

            # Collect failures
            failures = []

            # Read data starting from row 2 (assuming row 1 is header)
            max_row = sheet.max_row
            total_rows = max_row - 1  # Excluding header row

            # Create progress bar
            row_iterator = range(2, max_row + 1)
            if HAS_TQDM:
                row_iterator = tqdm(row_iterator, desc="Processing rows", unit="row", total=total_rows)

            for row_num in row_iterator:
                # Get Failure Count value (Column N)
                fcount_cell = sheet.cell(row=row_num, column=FCOUNT_COL)
                fcount = get_numeric_value(fcount_cell)

                # Only process rows with non-zero Failure Count
                if fcount != 0:
                    sn = get_cell_value(sheet.cell(row=row_num, column=SN_COL))
                    asic = get_cell_value(sheet.cell(row=row_num, column=ASIC_COL))
                    port = get_cell_value(sheet.cell(row=row_num, column=PORT_COL))
                    type_val = get_cell_value(sheet.cell(row=row_num, column=TYPE_COL))

                    failures.append({
                        'SN': sn,
                        'ASIC': asic,
                        'PORT': port,
                        'Failure Count': fcount,
                        'TYPE': type_val
                    })

            workbook.close()
        elif HAS_PANDAS:
            # Fallback to pandas method
            # Try to read the Excel file using pandas
            if HAS_TQDM:
                print("Reading Excel file...", file=sys.stderr)
            try:
                # Try with openpyxl engine first
                df = pd.read_excel(excel_path, engine='openpyxl', header=0)
            except (ImportError, ValueError):
                # If openpyxl not available, try default engine
                try:
                    df = pd.read_excel(excel_path, header=0)
                except Exception as e:
                    print(f"Error: Could not read Excel file with pandas: {e}", file=sys.stderr)
                    print("Please install openpyxl: pip install openpyxl", file=sys.stderr)
                    sys.exit(1)

            # Column indices (0-indexed in pandas)
            # Column C = index 2, Column J = index 9, Column K = index 10,
            # Column M = index 12, Column N = index 13
            columns = df.columns.tolist()

            if len(columns) < 14:
                print(f"Error: Expected at least 14 columns, found {len(columns)}", file=sys.stderr)
                sys.exit(1)

            # Get column values by index
            sn_col = columns[2]  # Column C
            asic_col = columns[9]  # Column J
            port_col = columns[10]  # Column K
            type_col = columns[12]  # Column M
            fcount_col = columns[13]  # Column N

            # Filter rows where Failure Count is non-zero
            if HAS_TQDM:
                print("Filtering rows...", file=sys.stderr)
            df[fcount_col] = pd.to_numeric(df[fcount_col], errors='coerce').fillna(0)
            failures_df = df[df[fcount_col] != 0]

            # Convert to list of dictionaries
            failures = []
            row_iterator = failures_df.iterrows()
            if HAS_TQDM:
                row_iterator = tqdm(row_iterator, desc="Processing failures", unit="row", total=len(failures_df))

            for idx, row in row_iterator:
                failures.append({
                    'SN': str(row[sn_col]) if pd.notna(row[sn_col]) else "",
                    'ASIC': str(row[asic_col]) if pd.notna(row[asic_col]) else "",
                    'PORT': str(row[port_col]) if pd.notna(row[port_col]) else "",
                    'Failure Count': int(row[fcount_col]),
                    'TYPE': str(row[type_col]) if pd.notna(row[type_col]) else ""
                })
        else:
            print("Error: Neither openpyxl nor pandas is available.", file=sys.stderr)
            print("Please install openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)

        if not failures:
            print("No failures found (all Failure Count values are zero).")
            return

        # Print header
        print(f"{'SN':<20} {'ASIC':<15} {'PORT':<10} {'Failure Count':<15} {'TYPE':<20}")
        print("-" * 85)

        # Print each failure row
        for failure in failures:
            print(f"{failure['SN']:<20} {failure['ASIC']:<15} {failure['PORT']:<10} "
                  f"{failure['Failure Count']:<15} {failure['TYPE']:<20}")

        print(f"\nTotal failures found: {len(failures)}")

    except Exception as e:
        print(f"Error reading Excel file: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    """Main entry point."""
    # Default file path
    default_file = Path(__file__).parent.parent / "QC3_S7TK_0128_build_test.xlsx"

    # Allow command-line argument for file path
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = default_file

    analyze_failures(excel_path)


if __name__ == "__main__":
    main()
