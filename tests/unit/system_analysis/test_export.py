"""Unit tests for Excel export module."""

from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import pytest
from openpyxl import load_workbook

from bh_glx_data.system_analysis.database import DatabaseStats
from bh_glx_data.system_analysis.export import ExcelExporter, ExcelExportResult
from bh_glx_data.system_analysis.query_engine import (
    AggregatedHostStats,
    BERHistogram,
    BERStatistics,
    CustomThresholdCounts,
    HostBERStats,
    LaneBERStats,
    ThresholdExceededCounts,
    TrainingFailureCounts,
)
from bh_glx_data.system_analysis.visualization import BER_COLOR_SCHEMES, ColorScheme


@pytest.fixture
def mock_db_manager():
    """Create mock DatabaseManager."""
    db = Mock()
    db.db_path = Path("/tmp/test.db")
    return db


@pytest.fixture
def exporter(mock_db_manager):
    """Create ExcelExporter instance with mock database."""
    return ExcelExporter(mock_db_manager)


@pytest.fixture
def sample_ber_statistics():
    """Create sample BER statistics."""
    lane_stats = {
        "01:00.0/ETH07/lane_0": LaneBERStats(
            lane_id="01:00.0/ETH07/lane_0",
            min_ber=1.23e-12,
            avg_ber=2.34e-11,
            max_ber=3.45e-10,
            high_ber_count=5,
            sample_count=100,
        ),
        "01:00.0/ETH07/lane_1": LaneBERStats(
            lane_id="01:00.0/ETH07/lane_1",
            min_ber=2.34e-12,
            avg_ber=3.45e-11,
            max_ber=4.56e-10,
            high_ber_count=3,
            sample_count=100,
        ),
    }
    return BERStatistics(
        lane_stats=lane_stats,
        num_tests=200,
        num_systems=2,
        train_speeds=[200],
    )


@pytest.fixture
def sample_threshold_counts():
    """Create sample threshold exceeded counts."""
    return ThresholdExceededCounts(
        lane_counts={
            "01:00.0/ETH07/lane_0": 10,
            "01:00.0/ETH07/lane_1": 5,
        },
        num_tests=15,
        num_systems=2,
        train_speeds=[200],
    )


@pytest.fixture
def sample_custom_threshold_counts():
    """Create sample custom threshold counts."""
    return CustomThresholdCounts(
        lane_counts={
            "01:00.0/ETH07/lane_0": 8,
            "01:00.0/ETH07/lane_1": 3,
        },
        num_tests=11,
        num_systems=2,
        threshold=1e-10,
        train_speeds=[200],
    )


@pytest.fixture
def sample_training_counts():
    """Create sample training failure counts."""
    return TrainingFailureCounts(
        lane_counts={
            "01:00.0/ETH07/lane_0": 12,
            "01:00.0/ETH07/lane_1": 7,
        },
        num_tests=19,
        num_systems=2,
        train_speeds=[200],
    )


@pytest.fixture
def sample_histogram():
    """Create sample BER histogram."""
    return BERHistogram(
        lane_id="01:00.0/ETH07/lane_0",
        bins=[
            ("< 1e-12", 50),
            ("1e-12 to 1e-11", 30),
            ("1e-11 to 1e-10", 15),
            ("> 1e-10", 5),
        ],
        num_tests=100,
        num_systems=2,
        train_speeds=[200],
    )


@pytest.fixture
def sample_aggregated_stats():
    """Create sample aggregated host statistics."""
    host_stats = [
        HostBERStats(
            host="system1",
            min_ber=1e-12,
            avg_ber=1e-11,
            max_ber=1e-10,
            sample_count=50,
        ),
        HostBERStats(
            host="system2",
            min_ber=2e-12,
            avg_ber=2e-11,
            max_ber=2e-10,
            sample_count=45,
        ),
    ]
    return AggregatedHostStats(
        lane_id="01:00.0/ETH07/lane_0",
        host_stats=host_stats,
        min_of_mins=1e-12,
        avg_of_mins=1.5e-12,
        max_of_mins=2e-12,
        min_of_avgs=1e-11,
        avg_of_avgs=1.5e-11,
        max_of_avgs=2e-11,
        min_of_maxs=1e-10,
        avg_of_maxs=1.5e-10,
        max_of_maxs=2e-10,
        num_systems=2,
        train_speeds=[200],
    )


@pytest.fixture
def sample_database_stats():
    """Create sample database statistics."""
    return DatabaseStats(
        total_tests=1000,
        unique_hosts=5,
        unique_speeds=[200, 400],
        date_range=("2026-01-01", "2026-03-16"),
        status_breakdown={"PASS": 900, "BER_THRESHOLD_EXCEEDED": 50, "TRAINING_FAIL": 50},
        total_ingestions=10,
    )


class TestExportBERStatistics:
    """Tests for export_ber_statistics method."""

    def test_export_table_format(self, exporter, sample_ber_statistics, tmp_path):
        """Test exporting BER statistics in table format."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all", format="table"
        )

        assert isinstance(result, ExcelExportResult)
        assert result.output_path == output_path
        assert result.worksheet_name == "Stats - table"
        assert result.rows_written == 2
        assert result.file_existed is False

        # Verify file was created
        assert output_path.exists()

        # Verify worksheet content
        wb = load_workbook(output_path)
        assert "Stats - table" in wb.sheetnames

        ws = wb["Stats - table"]
        # Header row (after title and spacing)
        assert ws.cell(3, 1).value == "bus_id"
        assert ws.cell(3, 2).value == "eth_id"
        assert ws.cell(3, 3).value == "lane"
        # Data row
        assert ws.cell(4, 1).value == "01:00.0"  # bus_id
        assert ws.cell(4, 2).value == "ETH07"  # eth_id
        assert ws.cell(4, 3).value == 0  # lane number (now numeric)

    def test_export_heatmap_format(self, exporter, sample_ber_statistics, tmp_path):
        """Test exporting BER statistics in heatmap format."""
        output_path = tmp_path / "test.xlsx"
        color_scheme = BER_COLOR_SCHEMES["default"]

        result = exporter.export_ber_statistics(
            sample_ber_statistics,
            output_path,
            lane_spec="01:00.0/ETH07",
            format="heatmap",
            color_scheme=color_scheme,
        )

        assert result.worksheet_name == "Stats - heatmap - max"
        assert result.rows_written == 2

        # Verify heatmap structure
        wb = load_workbook(output_path)
        ws = wb["Stats - heatmap - max"]
        assert ws.cell(1, 1).value  # Has header/port column

    def test_export_to_existing_file(self, exporter, sample_ber_statistics, tmp_path):
        """Test exporting to an existing Excel file."""
        output_path = tmp_path / "test.xlsx"

        # First export
        result1 = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all", format="table"
        )
        assert result1.file_existed is False

        # Second export to same file (same format, will create collision)
        result2 = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="01:00.0/*", format="table"
        )
        assert result2.file_existed is True

        # Verify both worksheets exist (both named "Stats - table" with collision numbering)
        wb = load_workbook(output_path)
        assert "Stats - table" in wb.sheetnames
        assert "Stats - table (2)" in wb.sheetnames

    def test_export_worksheet_name_collision(self, exporter, sample_ber_statistics, tmp_path):
        """Test handling worksheet name collisions."""
        output_path = tmp_path / "test.xlsx"

        # Export same format multiple times (all default to table)
        result1 = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all"
        )
        result2 = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all"
        )
        result3 = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all"
        )

        assert result1.worksheet_name == "Stats - table"
        assert result2.worksheet_name == "Stats - table (2)"
        assert result3.worksheet_name == "Stats - table (3)"


class TestExportCountData:
    """Tests for export_count_data method."""

    def test_export_threshold_counts(self, exporter, sample_threshold_counts, tmp_path):
        """Test exporting threshold exceeded counts."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_count_data(
            sample_threshold_counts, output_path, lane_spec="all", format="table"
        )

        assert result.worksheet_name == "Threshold - table"
        assert result.rows_written == 2

        wb = load_workbook(output_path)
        assert "Threshold - table" in wb.sheetnames

    def test_export_custom_threshold_counts(self, exporter, sample_custom_threshold_counts, tmp_path):
        """Test exporting custom threshold counts."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_count_data(
            sample_custom_threshold_counts, output_path, lane_spec="01:00.0/*", format="table"
        )

        assert result.worksheet_name == "Custom - table"

        # Verify threshold is included in metadata
        wb = load_workbook(output_path)
        ws = wb["Custom - table"]
        # Threshold should appear somewhere in the worksheet (as numeric value now)
        found_threshold = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, float) and abs(cell.value - 1e-10) < 1e-15:
                    found_threshold = True
                    break
            if found_threshold:
                break
        assert found_threshold

    def test_export_training_counts(self, exporter, sample_training_counts, tmp_path):
        """Test exporting training failure counts."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_count_data(
            sample_training_counts, output_path, lane_spec="all", format="table"
        )

        assert result.worksheet_name == "Training - table"

    def test_export_count_heatmap_format(self, exporter, sample_threshold_counts, tmp_path):
        """Test exporting counts in heatmap format."""
        output_path = tmp_path / "test.xlsx"
        color_scheme = BER_COLOR_SCHEMES["default"]

        result = exporter.export_count_data(
            sample_threshold_counts,
            output_path,
            lane_spec="all",
            format="heatmap",
            color_scheme=color_scheme,
        )

        assert result.rows_written == 2

        # Verify heatmap structure
        wb = load_workbook(output_path)
        ws = wb["Threshold - heatmap"]
        # Should have port column and lane columns
        assert ws.cell(1, 1).value  # Port column header


class TestExportHistogram:
    """Tests for export_histogram method."""

    def test_export_single_histogram(self, exporter, sample_histogram, tmp_path):
        """Test exporting single histogram."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_histogram(sample_histogram, output_path, lane_spec="01:00.0/ETH07/0")

        assert result.worksheet_name == "Histogram - chart"
        assert result.rows_written == 4  # 4 bins

        wb = load_workbook(output_path)
        ws = wb["Histogram - chart"]

        # Check histogram title (contains lane_spec)
        assert "01:00.0/ETH07/0" in str(ws.cell(1, 1).value)

        # Check bin data (row 5: title + spacing + hist title + headers + data)
        assert ws.cell(5, 1).value == "< 1e-12"
        assert ws.cell(5, 2).value == 50

    def test_export_multiple_histograms(self, exporter, tmp_path):
        """Test exporting multiple histograms."""
        output_path = tmp_path / "test.xlsx"

        histograms = [
            BERHistogram(
                lane_id="lane_0",
                bins=[("< 1e-12", 50)],
                num_tests=50,
                num_systems=2,
                train_speeds=[200],
            ),
            BERHistogram(
                lane_id="lane_1",
                bins=[("< 1e-12", 45)],
                num_tests=45,
                num_systems=2,
                train_speeds=[200],
            ),
        ]

        result = exporter.export_histogram(histograms, output_path, lane_spec="all")

        assert result.rows_written == 2  # 2 bins total

        wb = load_workbook(output_path)
        ws = wb["Histogram - chart"]

        # Both histograms should be present
        found_lane0 = False
        found_lane1 = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "lane_0" in str(cell.value):
                    found_lane0 = True
                if cell.value and "lane_1" in str(cell.value):
                    found_lane1 = True

        assert found_lane0
        assert found_lane1


class TestExportAdvancedStats:
    """Tests for export_advanced_stats method."""

    def test_export_single_lane_stats(self, exporter, sample_aggregated_stats, tmp_path):
        """Test exporting advanced stats for single lane."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_advanced_stats(
            sample_aggregated_stats, output_path, lane_spec="01:00.0/ETH07/0"
        )

        # New format doesn't include lane_spec
        assert result.worksheet_name == "Advanced Stats - table"
        # 2 host rows + 3 stats-of-stats rows
        assert result.rows_written == 5

        wb = load_workbook(output_path)
        ws = wb["Advanced Stats - table"]

        # Check per-host table
        assert "Per-Host Statistics" in str(ws.cell(3, 1).value)

        # Check host data
        assert "system1" in str(ws.cell(5, 1).value) or "system2" in str(ws.cell(5, 1).value)

        # Check stats-of-stats table
        stats_table_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Statistics of Host Statistics":
                    stats_table_found = True
                    break
            if stats_table_found:
                break
        assert stats_table_found

    def test_export_multiple_lane_stats(self, exporter, tmp_path):
        """Test exporting advanced stats for multiple lanes."""
        output_path = tmp_path / "test.xlsx"

        stats_list = [
            AggregatedHostStats(
                lane_id="lane_0",
                host_stats=[
                    HostBERStats(
                        host="sys1", min_ber=1e-12, avg_ber=1e-11, max_ber=1e-10, sample_count=50
                    )
                ],
                min_of_mins=1e-12,
                avg_of_mins=1e-12,
                max_of_mins=1e-12,
                min_of_avgs=1e-11,
                avg_of_avgs=1e-11,
                max_of_avgs=1e-11,
                min_of_maxs=1e-10,
                avg_of_maxs=1e-10,
                max_of_maxs=1e-10,
                num_systems=1,
                train_speeds=[200],
            ),
            AggregatedHostStats(
                lane_id="lane_1",
                host_stats=[
                    HostBERStats(
                        host="sys1", min_ber=2e-12, avg_ber=2e-11, max_ber=2e-10, sample_count=50
                    )
                ],
                min_of_mins=2e-12,
                avg_of_mins=2e-12,
                max_of_mins=2e-12,
                min_of_avgs=2e-11,
                avg_of_avgs=2e-11,
                max_of_avgs=2e-11,
                min_of_maxs=2e-10,
                avg_of_maxs=2e-10,
                max_of_maxs=2e-10,
                num_systems=1,
                train_speeds=[200],
            ),
        ]

        result = exporter.export_advanced_stats(stats_list, output_path, lane_spec="all")

        # 2 lanes × (1 host row + 3 stats rows) = 8 rows
        assert result.rows_written == 8


class TestExportDatabaseInfo:
    """Tests for export_database_info method."""

    def test_export_database_info(self, exporter, sample_database_stats, tmp_path):
        """Test exporting database information."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_database_info(sample_database_stats, output_path)

        assert result.worksheet_name == "Database Info"
        assert result.rows_written > 0

        wb = load_workbook(output_path)
        ws = wb["Database Info"]

        # Check database info appears
        assert ws.cell(3, 1).value == "Property"  # Header row (after title)

        # Check specific properties - now stored as numeric value
        found_total_tests = False
        for row in ws.iter_rows():
            if row[0].value == "Total Tests":
                found_total_tests = True
                # Value should be numeric (1000), not string ("1,000")
                assert row[1].value == 1000
                break

        assert found_total_tests

    def test_export_database_info_with_status_breakdown(
        self, exporter, sample_database_stats, tmp_path
    ):
        """Test database info includes status breakdown."""
        output_path = tmp_path / "test.xlsx"

        result = exporter.export_database_info(sample_database_stats, output_path)

        wb = load_workbook(output_path)
        ws = wb["Database Info"]

        # Check status breakdown appears
        found_pass = False
        for row in ws.iter_rows():
            if row[0].value and "PASS" in str(row[0].value):
                found_pass = True
                break

        assert found_pass


class TestExportErrorHandling:
    """Tests for export error handling."""

    def test_export_creates_parent_directories(self, exporter, sample_ber_statistics, tmp_path):
        """Test that export creates parent directories if they don't exist."""
        output_path = tmp_path / "subdir" / "nested" / "test.xlsx"

        result = exporter.export_ber_statistics(
            sample_ber_statistics, output_path, lane_spec="all"
        )

        assert output_path.exists()
        assert result.output_path == output_path
