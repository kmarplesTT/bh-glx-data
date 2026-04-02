"""Unit tests for Excel formatters module."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from bh_glx_data.system_analysis.excel_formatters import (
    apply_cell_background_color,
    create_or_open_workbook,
    generate_unique_worksheet_name,
    get_excel_color_for_value,
    map_terminal_color_to_excel,
    write_heatmap_to_worksheet,
    write_histogram_to_worksheet,
    write_metadata_section,
    write_table_to_worksheet,
)
from bh_glx_data.system_analysis.query_engine import BERHistogram
from bh_glx_data.system_analysis.visualization import BER_COLOR_SCHEMES, ColorScheme


class TestWorkbookManagement:
    """Tests for workbook creation and opening."""

    def test_create_new_workbook(self, tmp_path):
        """Test creating a new workbook."""
        file_path = tmp_path / "test.xlsx"
        wb, file_existed = create_or_open_workbook(file_path)

        assert isinstance(wb, Workbook)
        assert file_existed is False
        assert "Sheet" not in wb.sheetnames  # Default sheet removed

    def test_open_existing_workbook(self, tmp_path):
        """Test opening an existing workbook."""
        file_path = tmp_path / "test.xlsx"

        # Create initial workbook
        wb1 = Workbook()
        ws = wb1.create_sheet("TestSheet")
        ws.cell(1, 1, "Test Data")
        wb1.save(file_path)

        # Open existing workbook
        wb2, file_existed = create_or_open_workbook(file_path)

        assert isinstance(wb2, Workbook)
        assert file_existed is True
        assert "TestSheet" in wb2.sheetnames


class TestWorksheetNaming:
    """Tests for unique worksheet name generation."""

    def test_unique_name_no_collision(self):
        """Test generating unique name when no collision exists."""
        wb = Workbook()
        name = generate_unique_worksheet_name(wb, "Stats - all")

        assert name == "Stats - all"

    def test_unique_name_with_collision(self):
        """Test generating unique name when collision exists."""
        wb = Workbook()
        wb.create_sheet("Stats - all")

        name = generate_unique_worksheet_name(wb, "Stats - all")
        assert name == "Stats - all (2)"

    def test_unique_name_multiple_collisions(self):
        """Test generating unique name with multiple collisions."""
        wb = Workbook()
        wb.create_sheet("Stats - all")
        wb.create_sheet("Stats - all (2)")
        wb.create_sheet("Stats - all (3)")

        name = generate_unique_worksheet_name(wb, "Stats - all")
        assert name == "Stats - all (4)"

    def test_unique_name_truncation(self):
        """Test truncating long worksheet names."""
        wb = Workbook()
        long_name = "A" * 50  # Excel limit is 31 chars

        name = generate_unique_worksheet_name(wb, long_name)
        assert len(name) <= 31


class TestColorMapping:
    """Tests for color mapping functions."""

    def test_map_terminal_color_to_excel(self):
        """Test mapping terminal colors to Excel hex colors."""
        assert map_terminal_color_to_excel("color(28)") == "FF4CAF50"  # Green
        assert map_terminal_color_to_excel("color(106)") == "FF9DC34D"  # Yellow-green
        assert map_terminal_color_to_excel("color(184)") == "FFFFD700"  # Yellow
        assert map_terminal_color_to_excel("color(172)") == "FFFFA500"  # Orange
        assert map_terminal_color_to_excel("color(124)") == "FFDC143C"  # Red

    def test_map_terminal_color_unknown(self):
        """Test mapping unknown terminal color returns white."""
        assert map_terminal_color_to_excel("color(999)") == "FFFFFFFF"

    def test_get_excel_color_for_ber_value(self):
        """Test getting Excel color for BER values."""
        scheme = ColorScheme(
            thresholds=[
                (1e-12, "color(28)"),     # Green
                (1e-10, "color(106)"),    # Yellow-green
                (1e-8, "color(184)"),     # Yellow
                (1e-6, "color(172)"),     # Orange
            ],
            default_color="color(124)",  # Red
        )

        # Test BER values (lower is better)
        assert get_excel_color_for_value(1e-13, scheme, is_ber_metric=True) == "FF4CAF50"  # Green
        assert get_excel_color_for_value(5e-11, scheme, is_ber_metric=True) == "FF9DC34D"  # Yellow-green
        assert get_excel_color_for_value(5e-9, scheme, is_ber_metric=True) == "FFFFD700"  # Yellow
        assert get_excel_color_for_value(5e-7, scheme, is_ber_metric=True) == "FFFFA500"  # Orange
        assert get_excel_color_for_value(1e-5, scheme, is_ber_metric=True) == "FFDC143C"  # Red

    def test_get_excel_color_for_count_value(self):
        """Test getting Excel color for count values."""
        scheme = ColorScheme(
            thresholds=[
                (0, "color(28)"),       # Green for 0
                (10, "color(106)"),     # Yellow-green
                (50, "color(184)"),     # Yellow
                (100, "color(172)"),    # Orange
                (200, "color(124)"),    # Red
            ],
            default_color="color(124)",  # Red for above threshold
        )

        # Test count values (lower is better)
        assert get_excel_color_for_value(0, scheme, is_ber_metric=False) == "FF4CAF50"  # Green for 0
        assert get_excel_color_for_value(5, scheme, is_ber_metric=False) == "FF9DC34D"  # Yellow-green
        assert get_excel_color_for_value(30, scheme, is_ber_metric=False) == "FFFFD700"  # Yellow
        assert get_excel_color_for_value(75, scheme, is_ber_metric=False) == "FFFFA500"  # Orange
        assert get_excel_color_for_value(250, scheme, is_ber_metric=False) == "FFDC143C"  # Red

    def test_get_excel_color_for_none_value(self):
        """Test getting Excel color for None value returns white."""
        scheme = BER_COLOR_SCHEMES["default"]
        assert get_excel_color_for_value(None, scheme, is_ber_metric=True) == "FFFFFFFF"


class TestCellFormatting:
    """Tests for cell formatting functions."""

    def test_apply_cell_background_color(self):
        """Test applying background color to cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(1, 1, "Test")

        apply_cell_background_color(cell, "006400")

        assert cell.fill.start_color.rgb == "FF006400"  # openpyxl adds FF prefix

    def test_apply_cell_background_color_with_terminal_color(self):
        """Test applying background color using terminal color code."""
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(1, 1, "Test")

        apply_cell_background_color(cell, "color(28)")  # Green

        assert cell.fill.start_color.rgb == "FF4CAF50"


class TestMetadataSection:
    """Tests for metadata section writing."""

    def test_write_metadata_section(self):
        """Test writing metadata section to worksheet."""
        wb = Workbook()
        ws = wb.active

        metadata = {
            "Total Samples": 1000,
            "Unique Systems": 5,
            "Train Speeds": [200, 400],
        }

        next_row = write_metadata_section(ws, 1, metadata)

        # Check title
        assert ws.cell(1, 1).value == "Summary"
        assert ws.cell(1, 1).font.bold is True

        # Check metadata values (now stored as numbers, not formatted strings)
        assert ws.cell(2, 1).value == "Total Samples:"
        assert ws.cell(2, 2).value == 1000
        assert ws.cell(2, 2).number_format == "#,##0"

        assert ws.cell(3, 1).value == "Unique Systems:"
        assert ws.cell(3, 2).value == 5
        assert ws.cell(3, 2).number_format == "#,##0"

        assert ws.cell(4, 1).value == "Train Speeds:"
        assert ws.cell(4, 2).value == "200, 400"

        # Check next row is returned
        assert next_row == 6  # 4 data rows + 1 spacing


class TestTableWriting:
    """Tests for table writing function."""

    def test_write_table_to_worksheet_basic(self):
        """Test writing basic table to worksheet."""
        wb = Workbook()
        ws = wb.active

        data = {
            "Lane": ["lane_0", "lane_1", "lane_2"],
            "Min BER": ["1.23e-12", "2.34e-12", "3.45e-12"],
            "Max BER": ["5.67e-10", "6.78e-10", "7.89e-10"],
        }
        headers = ["Lane", "Min BER", "Max BER"]

        write_table_to_worksheet(ws, data, headers)

        # Check headers
        assert ws.cell(1, 1).value == "Lane"
        assert ws.cell(1, 2).value == "Min BER"
        assert ws.cell(1, 3).value == "Max BER"
        assert ws.cell(1, 1).font.bold is True

        # Check data
        assert ws.cell(2, 1).value == "lane_0"
        assert ws.cell(2, 2).value == "1.23e-12"
        assert ws.cell(3, 2).value == "2.34e-12"

    def test_write_table_with_title_and_metadata(self):
        """Test writing table with title and metadata."""
        wb = Workbook()
        ws = wb.active

        data = {"Lane": ["lane_0"], "Count": [10]}
        headers = ["Lane", "Count"]
        metadata = {"Total": 10}

        write_table_to_worksheet(ws, data, headers, title="Test Title", metadata=metadata)

        # Check title
        assert ws.cell(1, 1).value == "Test Title"
        assert ws.cell(1, 1).font.size == 14

        # Check headers (row 3 after title and spacing)
        assert ws.cell(3, 1).value == "Lane"

        # Check data
        assert ws.cell(4, 1).value == "lane_0"


class TestHeatmapWriting:
    """Tests for heatmap writing function."""

    def test_write_heatmap_to_worksheet_basic(self):
        """Test writing basic heatmap to worksheet."""
        wb = Workbook()
        ws = wb.active

        lane_data = {
            "01:00.0/ETH07/lane_0": 1.23e-12,
            "01:00.0/ETH07/lane_1": 2.34e-11,
            "01:00.0/ETH08/lane_0": 3.45e-10,
        }
        color_scheme = BER_COLOR_SCHEMES["default"]

        write_heatmap_to_worksheet(ws, lane_data, color_scheme, is_ber_metric=True)

        # Check header (now split into bus_id and eth_id)
        assert ws.cell(1, 1).value == "bus_id"
        assert ws.cell(1, 2).value == "eth_id"
        assert ws.cell(1, 3).value == "Lane 0"
        assert ws.cell(1, 4).value == "Lane 1"

        # Check port identifiers (now in separate columns)
        assert ws.cell(2, 1).value == "01:00.0"  # bus_id
        assert ws.cell(2, 2).value == "ETH07"    # eth_id
        assert ws.cell(3, 1).value == "01:00.0"  # bus_id
        assert ws.cell(3, 2).value == "ETH08"    # eth_id

        # Check lane values (now stored as float with scientific notation format)
        assert ws.cell(2, 3).value == 1.23e-12
        assert ws.cell(2, 3).number_format == "0.00E+00"

    def test_write_heatmap_with_colors(self):
        """Test that heatmap applies background colors."""
        wb = Workbook()
        ws = wb.active

        lane_data = {
            "01:00.0/ETH07/lane_0": 1e-13,  # Should be green
            "01:00.0/ETH07/lane_1": 1e-6,  # Should be red
        }
        color_scheme = ColorScheme(
            thresholds=[(1e-12, "color(28)"), (1e-7, "color(124)")],
            default_color="color(124)",
        )

        write_heatmap_to_worksheet(ws, lane_data, color_scheme, is_ber_metric=True)

        # Check colors are applied (green and red)
        assert ws.cell(2, 2).fill.start_color.rgb is not None
        assert ws.cell(2, 3).fill.start_color.rgb is not None

    def test_write_heatmap_with_metadata(self):
        """Test writing heatmap with metadata section."""
        wb = Workbook()
        ws = wb.active

        lane_data = {"01:00.0/ETH07/lane_0": 1e-12}
        color_scheme = BER_COLOR_SCHEMES["default"]
        metadata = {"Total Samples": 100}

        write_heatmap_to_worksheet(
            ws, lane_data, color_scheme, is_ber_metric=True, metadata=metadata
        )

        # Find metadata section (should be after heatmap data and legend)
        # Check that "Summary" appears somewhere in the worksheet
        found_summary = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Summary":
                    found_summary = True
                    break
            if found_summary:
                break

        assert found_summary


class TestHistogramWriting:
    """Tests for histogram writing function."""

    def test_write_histogram_to_worksheet_single(self):
        """Test writing single histogram to worksheet."""
        wb = Workbook()
        ws = wb.active

        histogram = BERHistogram(
            lane_id="01:00.0/ETH07/lane_0",
            bins=[
                ("< 1e-12", 50),
                ("1e-12 to 1e-11", 30),
                ("1e-11 to 1e-10", 15),
                ("> 1e-10", 5),
            ],
            num_tests=100,
            num_systems=2,
            train_speeds=[200],
        )

        write_histogram_to_worksheet(ws, histogram)

        # Check histogram title
        assert "01:00.0/ETH07/lane_0" in ws.cell(1, 1).value

        # Check data table headers
        assert ws.cell(2, 1).value == "BER Range"
        assert ws.cell(2, 2).value == "Count"

        # Check bin data
        assert ws.cell(3, 1).value == "< 1e-12"
        assert ws.cell(3, 2).value == 50
        assert ws.cell(4, 2).value == 30

    def test_write_histogram_with_metadata(self):
        """Test writing histogram with metadata section."""
        wb = Workbook()
        ws = wb.active

        histogram = BERHistogram(
            lane_id="01:00.0/ETH07/lane_0",
            bins=[("< 1e-12", 50)],
            num_tests=50,
            num_systems=2,
            train_speeds=[200],
        )
        metadata = {"Total Samples": 50}

        write_histogram_to_worksheet(ws, histogram, metadata=metadata)

        # Find metadata section
        found_summary = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Summary":
                    found_summary = True
                    break
            if found_summary:
                break

        assert found_summary

    def test_write_histogram_multiple(self):
        """Test writing multiple histograms to same worksheet."""
        wb = Workbook()
        ws = wb.active

        histograms = [
            BERHistogram(
                lane_id="lane_0",
                bins=[("< 1e-12", 50)],
                num_tests=50,
                num_systems=2,
                train_speeds=[200],
            ),
            BERHistogram(
                lane_id="lane_1",
                bins=[("< 1e-12", 45)],
                num_tests=45,
                num_systems=2,
                train_speeds=[200],
            ),
        ]

        write_histogram_to_worksheet(ws, histograms)

        # Check both histogram titles appear
        found_lane0 = False
        found_lane1 = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "lane_0" in str(cell.value):
                    found_lane0 = True
                if cell.value and "lane_1" in str(cell.value):
                    found_lane1 = True

        assert found_lane0
        assert found_lane1
