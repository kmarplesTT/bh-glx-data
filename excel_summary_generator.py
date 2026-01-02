"""
Excel Summary Generator for System Test Data

Processes CSV files from the data/ directory, groups them by system hostname and firmware version,
separates PRBS and Data test types, and generates Excel summary files using the
system_data_template.xlsx template.
"""
import sys
import re
import argparse
import logging
from pathlib import Path
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
DATA_DIR = Path('data')
TEMPLATE_PATH = Path('system_data_template.xlsx')
OUTPUT_DIR = Path('summaries')

# Test type constants
TEST_TYPE_PRBS = 'TestType.SERDES_PRBS'
TEST_TYPE_DATA = 'TestType.SIMPLE_PACKET'

# Sheet names in template
SHEET_RAW_PRBS = 'raw prbs data'
SHEET_RAW_DATA = 'raw data'
SHEET_PRBS_SUMMARY = 'PRBS Summary'
SHEET_DATA_SUMMARY = 'DATA Summary'


def scan_csv_files():
    """
    Scan data/ directory and return list of CSV files.

    Returns:
        list: List of Path objects for CSV files
    """
    if not DATA_DIR.exists():
        logger.error(f"Data directory {DATA_DIR} does not exist")
        return []

    csv_files = list(DATA_DIR.glob('*.csv'))
    logger.info(f"Found {len(csv_files)} CSV files in {DATA_DIR}")
    return csv_files


def extract_firmware_version(csv_path):
    """
    Extract firmware version from CSV filename.

    Looks for pattern like 'erisc_v1_7_103' or 'v1_7_103' in filename.

    Args:
        csv_path (Path): Path to CSV file

    Returns:
        str: Firmware version string (e.g., 'erisc_v1_7_103') or None if not found
    """
    filename = csv_path.name

    # Try to match erisc_vX_Y_Z pattern first
    pattern1 = r'erisc_v\d+_\d+_\d+'
    match = re.search(pattern1, filename)
    if match:
        return match.group(0)

    # Try to match vX_Y_Z pattern
    pattern2 = r'v\d+_\d+_\d+'
    match = re.search(pattern2, filename)
    if match:
        return match.group(0)

    logger.warning(f"Could not extract firmware version from filename: {filename}")
    return None


def identify_test_type(csv_path):
    """
    Determine if CSV is PRBS or Data test by reading the test_type column.

    Args:
        csv_path (Path): Path to CSV file

    Returns:
        str: 'PRBS' or 'DATA' or None if cannot determine
    """
    try:
        # Read just the first row to check test_type
        df = pd.read_csv(csv_path, nrows=1)

        if 'test_type' not in df.columns:
            logger.warning(f"CSV file {csv_path.name} does not have 'test_type' column")
            # Fallback to filename check
            if 'prbs_test' in csv_path.name.lower():
                return 'PRBS'
            elif 'data_test' in csv_path.name.lower():
                return 'DATA'
            return None

        test_type_value = df['test_type'].iloc[0]

        if test_type_value == TEST_TYPE_PRBS:
            return 'PRBS'
        elif test_type_value == TEST_TYPE_DATA:
            return 'DATA'
        else:
            logger.warning(f"Unknown test_type '{test_type_value}' in {csv_path.name}")
            # Fallback to filename check
            if 'prbs_test' in csv_path.name.lower():
                return 'PRBS'
            elif 'data_test' in csv_path.name.lower():
                return 'DATA'
            return None

    except Exception as e:
        logger.error(f"Error reading CSV file {csv_path.name}: {e}")
        # Fallback to filename check
        if 'prbs_test' in csv_path.name.lower():
            return 'PRBS'
        elif 'data_test' in csv_path.name.lower():
            return 'DATA'
        return None


def extract_system_hostname(csv_path):
    """
    Extract system hostname from CSV host column.

    Args:
        csv_path (Path): Path to CSV file

    Returns:
        str: System hostname (e.g., 'bh-glx-b02u02') or None if not found
    """
    try:
        # Read just the first row to get hostname
        df = pd.read_csv(csv_path, nrows=1)

        if 'host' not in df.columns:
            logger.warning(f"CSV file {csv_path.name} does not have 'host' column")
            return None

        hostname = df['host'].iloc[0]
        if pd.isna(hostname) or hostname == '':
            logger.warning(f"Empty hostname in {csv_path.name}")
            return None

        return str(hostname).strip()

    except Exception as e:
        logger.error(f"Error extracting hostname from {csv_path.name}: {e}")
        return None


def group_csvs_by_system_and_firmware(csv_files):
    """
    Group CSV files by system hostname AND firmware version.

    Args:
        csv_files (list): List of Path objects for CSV files

    Returns:
        dict: Dictionary with keys (hostname, firmware_version) and values as lists of CSV paths
    """
    grouped = defaultdict(lambda: {'PRBS': [], 'DATA': []})

    for csv_path in csv_files:
        hostname = extract_system_hostname(csv_path)
        firmware_version = extract_firmware_version(csv_path)
        test_type = identify_test_type(csv_path)

        if not hostname:
            logger.warning(f"Skipping {csv_path.name}: could not extract hostname")
            continue

        if not firmware_version:
            logger.warning(f"Skipping {csv_path.name}: could not extract firmware version")
            continue

        if not test_type:
            logger.warning(f"Skipping {csv_path.name}: could not identify test type")
            continue

        key = (hostname, firmware_version)
        grouped[key][test_type].append(csv_path)
        logger.debug(f"Grouped {csv_path.name}: hostname={hostname}, firmware={firmware_version}, type={test_type}")

    return grouped


def compile_test_data(csv_files, test_type):
    """
    Combine CSV data for a test type by reading and concatenating all CSV files.

    Args:
        csv_files (list): List of Path objects for CSV files of the same test type
        test_type (str): 'PRBS' or 'DATA'

    Returns:
        pd.DataFrame: Combined dataframe with all rows from all CSV files, or None if error
    """
    if not csv_files:
        logger.warning(f"No CSV files provided for {test_type} test type")
        return None

    dataframes = []

    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path)
            if df.empty:
                logger.warning(f"CSV file {csv_path.name} is empty")
                continue
            dataframes.append(df)
            logger.debug(f"Loaded {len(df)} rows from {csv_path.name}")
        except Exception as e:
            logger.error(f"Error reading CSV file {csv_path.name}: {e}")
            continue

    if not dataframes:
        logger.warning(f"No valid data found for {test_type} test type")
        return None

    # Concatenate all dataframes
    combined_df = pd.concat(dataframes, ignore_index=True)
    logger.info(f"Compiled {len(combined_df)} total rows for {test_type} test type from {len(dataframes)} files")

    return combined_df


def load_template():
    """
    Load system_data_template.xlsx workbook.

    Returns:
        openpyxl.Workbook: Workbook object or None if error
    """
    if not TEMPLATE_PATH.exists():
        logger.error(f"Template file {TEMPLATE_PATH} does not exist")
        return None

    try:
        workbook = openpyxl.load_workbook(TEMPLATE_PATH)
        logger.info(f"Loaded template from {TEMPLATE_PATH}")
        return workbook
    except Exception as e:
        logger.error(f"Error loading template {TEMPLATE_PATH}: {e}")
        return None


def paste_data_to_sheet(workbook, sheet_name, data_df):
    """
    Paste compiled data to specified sheet, starting from row 1.
    Overwrites existing data in the sheet.

    Args:
        workbook (openpyxl.Workbook): Workbook object
        sheet_name (str): Name of the sheet to paste data into
        data_df (pd.DataFrame): DataFrame to paste

    Returns:
        tuple: (last_row, last_col) tuple with 1-based indices, or (None, None) if error
    """
    if sheet_name not in workbook.sheetnames:
        logger.error(f"Sheet '{sheet_name}' not found in workbook")
        return None, None

    try:
        sheet = workbook[sheet_name]

        # Clear existing data if there is any
        if sheet.max_row > 0:
            # Delete all rows from 1 to max_row
            sheet.delete_rows(1, sheet.max_row)

        # Write headers
        headers = list(data_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = str(header) if header is not None else ''

        # Write data rows
        for row_idx, (_, row_data) in enumerate(data_df.iterrows(), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # Handle NaN, None, and other special values
                if pd.isna(value):
                    cell.value = None
                elif value is None:
                    cell.value = None
                else:
                    # Convert to string for complex objects, otherwise keep as-is
                    try:
                        cell.value = value
                    except (TypeError, ValueError):
                        cell.value = str(value)

        last_row = len(data_df) + 1  # +1 for header row
        last_col = len(headers)

        logger.info(f"Pasted {len(data_df)} rows to sheet '{sheet_name}' (range: A1:{get_column_letter(last_col)}{last_row})")
        return last_row, last_col

    except Exception as e:
        logger.error(f"Error pasting data to sheet '{sheet_name}': {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None, None


def update_pivot_table_source(workbook, pivot_sheet_name, data_sheet_name, data_range):
    """
    Update pivot table data source in the specified sheet.

    Args:
        workbook (openpyxl.Workbook): Workbook object
        pivot_sheet_name (str): Name of the sheet containing the pivot table
        data_sheet_name (str): Name of the sheet containing the data
        data_range (str): Excel range string (e.g., 'A1:Z1000')

    Returns:
        bool: True if successful, False otherwise
    """
    if pivot_sheet_name not in workbook.sheetnames:
        logger.error(f"Pivot sheet '{pivot_sheet_name}' not found in workbook")
        return False

    try:
        sheet = workbook[pivot_sheet_name]

        # Access pivot tables through the sheet
        # In openpyxl, pivot tables are accessed via sheet._pivots (private attribute)
        if not hasattr(sheet, '_pivots') or not sheet._pivots:
            logger.warning(f"No pivot tables found in sheet '{pivot_sheet_name}'")
            return False

        # Update each pivot table's data source
        updated_count = 0
        for pivot_table in sheet._pivots:
            try:
                # Access the cache
                cache = pivot_table.cache
                if cache and hasattr(cache, 'cacheSource'):
                    source = cache.cacheSource
                    if hasattr(source, 'worksheetSource'):
                        ws_source = source.worksheetSource
                        # Update the range and sheet name
                        ws_source.ref = data_range
                        ws_source.sheet = data_sheet_name
                        updated_count += 1
                        logger.info(f"Updated pivot table data source to '{data_sheet_name}'!{data_range}")
            except Exception as e:
                logger.warning(f"Could not update one pivot table: {e}")
                continue

        if updated_count > 0:
            return True
        else:
            logger.warning(f"Could not update any pivot tables in sheet '{pivot_sheet_name}'")
            return False

    except Exception as e:
        logger.error(f"Error updating pivot table source in sheet '{pivot_sheet_name}': {e}")
        return False


def generate_excel_summary(hostname, firmware_version, prbs_data, data_test_data):
    """
    Main function to generate Excel file for a system+firmware combination.

    Args:
        hostname (str): System hostname (e.g., 'bh-glx-b02u02')
        firmware_version (str): Firmware version (e.g., 'erisc_v1_7_103')
        prbs_data (pd.DataFrame): PRBS test data or None
        data_test_data (pd.DataFrame): Data test data or None

    Returns:
        bool: True if successful, False otherwise
    """
    # Load template
    workbook = load_template()
    if not workbook:
        return False

    # Create output directory if it doesn't exist
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Generate output filename
    output_filename = f"{hostname}_{firmware_version}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    # Paste PRBS data if available
    prbs_range = None
    if prbs_data is not None and not prbs_data.empty:
        last_row, last_col = paste_data_to_sheet(workbook, SHEET_RAW_PRBS, prbs_data)
        if last_row and last_col:
            prbs_range = f"A1:{get_column_letter(last_col)}{last_row}"
            # Update PRBS Summary pivot table if sheet exists
            if SHEET_PRBS_SUMMARY in workbook.sheetnames:
                update_pivot_table_source(workbook, SHEET_PRBS_SUMMARY, SHEET_RAW_PRBS, prbs_range)
            else:
                logger.warning(f"Sheet '{SHEET_PRBS_SUMMARY}' not found, skipping pivot table update")
        else:
            logger.warning(f"Failed to paste PRBS data for {hostname} {firmware_version}")
    else:
        logger.info(f"No PRBS data for {hostname} {firmware_version}")

    # Paste Data test data if available
    data_range = None
    if data_test_data is not None and not data_test_data.empty:
        last_row, last_col = paste_data_to_sheet(workbook, SHEET_RAW_DATA, data_test_data)
        if last_row and last_col:
            data_range = f"A1:{get_column_letter(last_col)}{last_row}"
            # Update DATA Summary pivot table if sheet exists
            if SHEET_DATA_SUMMARY in workbook.sheetnames:
                update_pivot_table_source(workbook, SHEET_DATA_SUMMARY, SHEET_RAW_DATA, data_range)
            else:
                logger.warning(f"Sheet '{SHEET_DATA_SUMMARY}' not found, skipping pivot table update")
        else:
            logger.warning(f"Failed to paste Data test data for {hostname} {firmware_version}")
    else:
        logger.info(f"No Data test data for {hostname} {firmware_version}")

    # Save workbook
    try:
        workbook.save(output_path)
        logger.info(f"Saved Excel summary to {output_path}")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel file {output_path}: {e}")
        return False


def parse_arguments():
    """
    Parse command-line arguments.

    Returns:
        argparse.Namespace: Parsed arguments
    """
    parser = argparse.ArgumentParser(
        description='Generate Excel summaries from CSV test data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                          # Process all systems
  %(prog)s --systems bh-glx-b02u02   # Process a single system
  %(prog)s --systems bh-glx-b02u02 bh-glx-b03u02  # Process multiple systems
        """
    )
    parser.add_argument(
        '--systems',
        nargs='*',
        help='System hostname(s) to process (e.g., bh-glx-b02u02). If not provided, processes all systems'
    )
    return parser.parse_args()


def main():
    """
    Main entry point for the script.
    """
    # Parse command-line arguments
    args = parse_arguments()

    logger.info("Starting Excel summary generation")

    # Scan CSV files
    csv_files = scan_csv_files()
    if not csv_files:
        logger.error("No CSV files found. Exiting.")
        sys.exit(1)

    # Group CSV files by system and firmware
    grouped = group_csvs_by_system_and_firmware(csv_files)

    if not grouped:
        logger.error("No valid CSV files could be grouped. Exiting.")
        sys.exit(1)

    # Filter by systems if specified
    if args.systems:
        systems_to_process = set(args.systems)
        filtered_grouped = {
            (hostname, firmware_version): file_groups
            for (hostname, firmware_version), file_groups in grouped.items()
            if hostname in systems_to_process
        }

        if not filtered_grouped:
            logger.error(f"No data found for specified systems: {', '.join(args.systems)}")
            sys.exit(1)

        # Check if any requested systems were not found
        found_systems = {hostname for hostname, _ in filtered_grouped.keys()}
        missing_systems = systems_to_process - found_systems
        if missing_systems:
            logger.warning(f"Warning: No data found for systems: {', '.join(missing_systems)}")

        grouped = filtered_grouped
        logger.info(f"Filtering to {len(grouped)} system+firmware combinations for specified systems")
    else:
        logger.info(f"Processing all {len(grouped)} unique system+firmware combinations")

    # Process each system+firmware combination
    success_count = 0
    error_count = 0

    for (hostname, firmware_version), file_groups in grouped.items():
        logger.info(f"\nProcessing {hostname} with firmware {firmware_version}")

        # Compile PRBS data
        prbs_data = None
        if file_groups['PRBS']:
            prbs_data = compile_test_data(file_groups['PRBS'], 'PRBS')

        # Compile Data test data
        data_test_data = None
        if file_groups['DATA']:
            data_test_data = compile_test_data(file_groups['DATA'], 'DATA')

        # Generate Excel file
        if prbs_data is None and data_test_data is None:
            logger.warning(f"No data available for {hostname} {firmware_version}, skipping")
            error_count += 1
            continue

        if generate_excel_summary(hostname, firmware_version, prbs_data, data_test_data):
            success_count += 1
        else:
            error_count += 1

    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"System+firmware combinations processed: {len(grouped)}")
    print(f"Successfully generated: {success_count}")
    print(f"Errors: {error_count}")
    print(f"Output directory: {OUTPUT_DIR.absolute()}")


if __name__ == '__main__':
    main()
